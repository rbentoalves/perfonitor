import pandas as pd
from datetime import datetime
import re
## import PySimpleGUI as sg
import FreeSimpleGUI as sg
import math
import inputs as inputs
import os
import openpyxl
import sys


# <editor-fold desc="Corrections">

def correct_incidents_irradiance_for_overlapping_parents(incidents, irradiance, export, component_data,
                                                         recalculate: bool = False, timestamp: float = 15,
                                                         irradiance_threshold: float = 20):
    """From: Incidents table, irradiance
    Returns: Dict with Irradiance dataframe corrected for overlapping parents events, i.e.,
    removes periods where parents incidents are active. For each incident"""

    incidents_corrected_info = {}
    granularity = timestamp / 60
    granularity_str = str(timestamp) + "min"

    if recalculate is True:
        incidents_to_correct = incidents
        pass
    else:
        n_inc_1 = incidents.shape[0]
        incidents_to_correct = incidents.loc[(incidents['Event End Time'].isna()) |
                                             (incidents['Energy Lost (MWh)'].isna())]

        n_inc_2 = incidents_to_correct.shape[0]
        print('No recalculation, analysing overlappers on ', n_inc_2, ' from a total of ', n_inc_1, ' incidents.')

    for index, row in incidents_to_correct.iterrows():

        # Get site info
        site = row['Site Name']
        site_info = component_data.loc[component_data['Site'] == site]
        site_capacity = float(component_data.loc[component_data['Component'] == site]['Nominal Power DC'].values)
        'budget_pr_site = budget_pr.loc[site, :]'

        # Get site Incidents
        site_incidents = incidents.loc[incidents['Site Name'] == site]

        # Get site irradiance
        df_irradiance_site = irradiance.loc[:, irradiance.columns.str.contains(site + '|Timestamp')]
        df_export_site = export.loc[:, export.columns.str.contains(site + '|Timestamp')]

        # Get irradiance poa avg column and export column
        poa_avg_column = df_irradiance_site.loc[:, df_irradiance_site.columns.str.contains('Irradiance')]
        poa_avg_column = poa_avg_column.loc[:, ~poa_avg_column.columns.str.contains('curated')].columns.values[0]

        export_column = df_export_site.loc[:, df_export_site.columns.str.contains(site)].columns.values[0]

        # Incident Info
        id_incident = row['ID']
        capacity = row['Capacity Related Component']
        parents = (site_info.loc[site_info['Component'] == row['Related Component']]).loc[:,
                  site_info.columns.str.contains('Parent')].values.flatten().tolist()
        parents = [x for x in parents if str(x) != 'nan']
        print("\n" + id_incident)

        try:
            if pd.isnull(row['Event End Time']):
                # Active Events-------------------------------------------------------------------------------------
                effective_start_time_incident = row['Event Start Time']

                # In active events, end time of incident is the latest record of irradiance
                effective_end_time_incident = datetime.strptime(str(df_irradiance_site['Timestamp'].to_list()[-1]),
                                                                '%Y-%m-%d %H:%M:%S')
                print("Active incident, effective end time: " + str(effective_end_time_incident))
                closed_event = False

            else:
                # Closed Events
                closed_event = True
                effective_start_time_incident = row['Event Start Time']
                effective_end_time_incident = row['Event End Time']
                print("Closed incident")

        except TypeError:
            # Closed Events
            closed_event = True
            effective_start_time_incident = row['Event Start Time']
            effective_end_time_incident = row['Event End Time']
            print("Closed incident")

        finally:
            if len(parents) == 0:
                pass
            else:
                parents_incidents = site_incidents[site_incidents['Related Component'].isin(parents)]
                if not parents_incidents.empty:
                    relevant_parents_incidents = parents_incidents.loc[
                        ~(parents_incidents['Event End Time'] <= effective_start_time_incident) & ~(
                            parents_incidents['Event Start Time'] >= effective_end_time_incident)]

                    # relevant_parents_incidents = mf.rounddatesclosed_15m(site,relevant_parents_incidents)

                    if not relevant_parents_incidents.empty:
                        # cycle through parents incidents and get active hours to remove
                        # Get first timestamp under analysis and df from that timestamp onwards
                        irradiance_incident = df_irradiance_site.loc[
                            (df_irradiance_site['Timestamp'] >= effective_start_time_incident) & (
                                df_irradiance_site['Timestamp'] <= effective_end_time_incident)]

                        irradiance_incident['Day'] = [datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S').date() for
                                                      timestamp in irradiance_incident['Timestamp']]

                        export_incident = df_export_site.loc[
                            (df_export_site['Timestamp'] >= effective_start_time_incident) & (
                                df_export_site['Timestamp'] <= effective_end_time_incident)]

                        export_incident['Day'] = [datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S').date() for
                                                  timestamp in export_incident['Timestamp']]

                        try:
                            del timestamps_to_remove
                            del timestamps_to_remove_eloss
                        except NameError:
                            pass

                        if closed_event is False:
                            print('Relevant overlapping Incidents for an active event')
                        else:
                            print('Relevant overlapping Incidents for a closed event')

                        print(row[['Related Component', 'Event Start Time', 'Event End Time']])
                        print(relevant_parents_incidents[['Related Component', 'Event Start Time', 'Event End Time']])

                        for index_rpi, row_rpi in relevant_parents_incidents.iterrows():
                            failure_mode = row_rpi["Failure Mode"]

                            rpi_start_time = \
                                pd.Series(row_rpi['Event Start Time']).dt.round(granularity_str, 'shift_backward')[0]
                            rpi_actual_start_time = row_rpi['Event Start Time']

                            if not pd.isnull(row_rpi["Event End Time"]):
                                print("End time not null: " + str(row_rpi["Event End Time"]))
                                try:
                                    rpi_end_time = \
                                        pd.Series(row_rpi['Event End Time']).dt.round(granularity_str, 'shift_forward')[
                                            0]
                                    rpi_actual_end_time = row_rpi['Event End Time']
                                    # print(rpi_actual_end_time)

                                except AttributeError:
                                    rpi_end_time = effective_end_time_incident
                                    rpi_actual_end_time = effective_end_time_incident
                            else:
                                print("End time null: " + str(row_rpi["Event End Time"]))
                                rpi_end_time = effective_end_time_incident
                                rpi_actual_end_time = effective_end_time_incident
                                print("New End time: " + str(effective_end_time_incident))

                            actual_timestamp_range = list(pd.date_range(start=rpi_start_time,
                                                                        end=rpi_end_time, freq='1min'))

                            if row_rpi["Failure Mode"] == "Curtailment":
                                print("Correcting for curtailment")
                                #print(type(rpi_actual_end_time))
                                #print(type(rpi_actual_start_time))

                                timestamp_range = export_incident.loc[
                                    (export_incident["Timestamp"] <= rpi_actual_end_time) &
                                    (export_incident["Timestamp"] >= rpi_actual_start_time) &
                                    (export_incident[export_column] <= 0)]["Timestamp"].to_list()  # for availability calculation (active incident hours)

                                timestamp_range_eloss = []  # this is empty bc curt excludes the capacity of the
                                # incident so energy loss needs to be calculated at this incident
                                # level

                                if len(timestamp_range) > 0:
                                    print("Between: " + str(timestamp_range[0]) + " and " + str(timestamp_range[-1]))
                                else:
                                    print("No correction needed")

                            else:
                                print("Correcting for non-curtailment")
                                timestamp_range = timestamp_range_eloss = list(pd.date_range(start=rpi_start_time,
                                                                                             end=rpi_end_time,
                                                                                             freq='15min'))

                                #print("Between: " + str(timestamp_range[0]) + " and " + str(timestamp_range[-1]))

                            try:
                                timestamps_to_remove += timestamp_range
                                actual_timestamps_to_remove += actual_timestamp_range
                                timestamps_to_remove_eloss += timestamp_range_eloss

                            except (NameError, AttributeError):
                                timestamps_to_remove = timestamp_range
                                actual_timestamps_to_remove = actual_timestamp_range
                                timestamps_to_remove_eloss = timestamp_range_eloss

                            # print("Look Here: \n", timestamps_to_remove)

                        timestamps_to_remove = sorted(set(timestamps_to_remove))
                        actual_timestamps_to_remove = sorted(set(actual_timestamps_to_remove))
                        timestamps_to_remove_eloss = sorted(set(timestamps_to_remove_eloss))

                        overlapped_time_1m = len(actual_timestamps_to_remove) / 60

                        #timetamps to keep for active hours
                        timestamps_to_keep = [timestamp for timestamp in irradiance_incident['Timestamp'].to_list() if
                                              timestamp not in timestamps_to_remove]

                        # timetamps to keep for energy loss
                        timestamps_to_keep_eloss = [timestamp for timestamp in
                                                    irradiance_incident['Timestamp'].to_list() if timestamp
                                                    not in timestamps_to_remove_eloss]

                        # irradiance for active hours
                        corrected_irradiance_incident = irradiance_incident.loc[
                            irradiance_incident['Timestamp'].isin(timestamps_to_keep)]

                        actual_column = corrected_irradiance_incident.loc[:,
                                        corrected_irradiance_incident.columns.str.contains('Average')]
                        actual_column = \
                            actual_column.loc[:, ~actual_column.columns.str.contains('curated')].columns.values[0]

                        # irradiance for energy lost
                        cleaned_irradiance = irradiance_incident.loc[
                            irradiance_incident['Timestamp'].isin(timestamps_to_keep_eloss)].dropna(
                            subset=[actual_column])

                        try:
                            data_gaps_proportion = 1 - (len(cleaned_irradiance[actual_column]) / len(
                                corrected_irradiance_incident[actual_column]))
                        except ZeroDivisionError:
                            print('Division by zero: ', len(cleaned_irradiance[actual_column]), " /",
                                  len(corrected_irradiance_incident[actual_column]))
                            data_gaps_proportion = 1

                        incidents_corrected_info[id_incident] = {
                            'Corrected Irradiance Incident': corrected_irradiance_incident,
                            'Cleaned Corrected Irradiance Incident': cleaned_irradiance,
                            "Time overlapped 1m": overlapped_time_1m,
                            'Data Gaps Proportion': data_gaps_proportion,
                            'Irradiance Column': actual_column,
                            'Irradiance Raw': irradiance_incident}

                        # print(irradiance_incident)
                    else:
                        print('No overlapping parents')
                        continue

                else:
                    print('No overlapping parents')
                    continue

    return incidents_corrected_info


def correct_duration_of_event(df):
    for index, row in df.iterrows():
        stime = df.loc[index, 'Event Start Time']
        stime = datetime.strptime(stime, '%Y-%m-%d %H:%M:%S')

        etime = df.loc[index, 'Event End Time']
        etime = datetime.strptime(etime, '%Y-%m-%d %H:%M:%S')

        duration = etime - stime
        duration = duration.total_seconds() / 3600

        df.loc[index, 'Duration (h)'] = round(duration, 2)

    return df


def correct_site_name(site):
    while site[-1] == " ":
        site = site[:-1]

    while site[0] == " ":
        site = site[1:]

    return site


def filter_notprod_and_duration(df, duration):
    ''' Filters df by not producing events and given duration in minutes

     Returns df'''
    d_hours = duration / 60.1  # turns minutes into hours

    df_closed = df.loc[df['Event End Time'].notnull()]

    df_15m_notprod = df_closed.loc[
        df['Component Status'] == 'Not Producing']  # new dataframe with only the not producing incidents

    df_15m_notprod_dur15m = df_15m_notprod.loc[
        df_15m_notprod['Duration (h)'] > d_hours]  # filter by duration >X minutes

    df_final = df_15m_notprod_dur15m.reset_index(None, drop=True)  # resets index

    return df_final


def get_actual_irradiance_column(df_irradiance_site):
    ''' From irradiance data of a site, containing the different types (curated, poa avg) selects curated if data gaps
    are less than 25%, if not it selects poa average if data gaps for that is less than 60%, if data gaps persist,
    returns no irradiance given that the results are invalid'''

    # Search for irradiance columns to use, if there is no curated irradiance, use POA average
    try:
        curated_column = df_irradiance_site.loc[:, df_irradiance_site.columns.str.contains('curated')].columns.values[
            0]
    except IndexError:
        curated_column = None
        print("no curated column")
    try:
        poa_avg_column = df_irradiance_site.loc[:, df_irradiance_site.columns.str.contains('Average')]
        poa_avg_column = poa_avg_column.loc[:, ~poa_avg_column.columns.str.contains('curated')].columns.values[0]
    except IndexError:
        poa_avg_column = None

    # print(curated_column)
    # print(poa_avg_column)

    # Data column check to determine what irradiance to use
    if curated_column:
        # Verify data gap percentage of curated column & then POA average
        cleaned_irradiance = [value for value in df_irradiance_site[curated_column] if not math.isnan(value)]
        # print(len(cleaned_irradiance),len(df_irradiance_site[curated_column]))
        try:
            data_gaps_proportion = 1 - (len(cleaned_irradiance) / len(df_irradiance_site[curated_column]))
        except ZeroDivisionError:
            data_gaps_proportion = 1

        if data_gaps_proportion > 0.1:

            # verify POA average
            try:
                cleaned_irradiance = [value for value in df_irradiance_site[poa_avg_column] if not math.isnan(value)]
                data_gaps_proportion = 1 - (len(cleaned_irradiance) / len(df_irradiance_site[poa_avg_column]))
            except ZeroDivisionError:
                data_gaps_proportion = 1

            if data_gaps_proportion > 0.60:
                actual_column = None

            else:
                actual_column = poa_avg_column


        else:
            actual_column = curated_column



    elif poa_avg_column:
        actual_column = poa_avg_column

        # Verify data gap percentage of poa avg column
        try:
            cleaned_irradiance = [value for value in df_irradiance_site[actual_column] if not math.isnan(value)]
            data_gaps_proportion = 1 - (len(cleaned_irradiance) / len(df_irradiance_site[actual_column]))

            if data_gaps_proportion > 0.60:
                actual_column = None
        except ZeroDivisionError:
            actual_column = None
            data_gaps_proportion = 1

    else:
        actual_column = None
        data_gaps_proportion = 1

    return actual_column, curated_column, data_gaps_proportion, poa_avg_column


def rename_dict_keys(d, keys):
    return dict([(keys.get(k), v) for k, v in d.items()])


def reset_final_report(Report_template_path, date, geography):
    year = date[:4]
    month = date[5:7]
    day = date[-2:]
    reportxl = openpyxl.load_workbook(Report_template_path)
    dir = os.path.dirname(Report_template_path)
    dir = dir.replace("/Info&Templates", "")
    basename = os.path.basename(Report_template_path)

    reportfile = dir + '/Reporting_' + geography + '_Sites_' + str(day) + '-' + str(month) + '.xlsx'

    if not 'Active Events' in reportxl.sheetnames:
        pass
        print('Active Events sheet does not exist')
        reportxl.save(reportfile)
    elif reportxl['Active Events'].max_row == 1:
        pass
        print('Active Events sheet is empty')
        reportxl.save(reportfile)
    else:
        del reportxl['Active Events']
        reportxl.save(reportfile)
    # -----------------------------------------------------------
    if not 'Closed Events' in reportxl.sheetnames:
        pass
        print('Closed Events sheet does not exist')
        reportxl.save(reportfile)
    elif reportxl['Closed Events'].max_row == 1:
        pass
        print('Closed Events sheet is empty')
        reportxl.save(reportfile)
    else:
        del reportxl['Closed Events']
        reportxl.save(reportfile)
    # -----------------------------------------------------------
    if not 'Active tracker incidents' in reportxl.sheetnames:
        pass
        print('Active tracker events sheet does not exist')
        reportxl.save(reportfile)
    elif reportxl['Active tracker incidents'].max_row == 1:
        pass
        print('Active tracker events sheet is empty')
        reportxl.save(reportfile)
    else:
        del reportxl['Active tracker incidents']
        reportxl.save(reportfile)
    # -----------------------------------------------------------
    if not 'Closed tracker incidents' in reportxl.sheetnames:
        pass
        print('Closed tracker events sheet does not exist')
        reportxl.save(reportfile)
    elif reportxl['Closed tracker incidents'].max_row == 1:
        pass
        print('Closed tracker events sheet is empty')
        reportxl.save(reportfile)
    else:
        del reportxl['Closed tracker incidents']
        reportxl.save(reportfile)

    return reportfile


# noinspection PyTypeChecker
def get_site_and_inverter_data_ready_for_analysis(inverters_data, component_data, budget_pr, general_info):
    inverter_list_raw = inverters_data.columns[inverters_data.columns.str.contains('AC')].to_list()
    inverter_list = [re.search(r'Inverter \d.+', name).group().replace(']', "") for name in inverter_list_raw]

    irradiance_columns = inverters_data.columns[inverters_data.columns.str.contains('Irradiance')].to_list()
    for irradiance in irradiance_columns:
        if 'curated' in irradiance:
            irradiance_curated_column = irradiance
        else:
            irradiance_poaavg_column = irradiance

    # print(irradiance_curated_column)

    site_list = list(set(
        [re.search(r'\[\w.+', name).group().replace(']', "").replace('[', "") for name in irradiance_columns]))

    if len(site_list) > 1:
        print('More than one site in irradiance data')
        exit()
    else:
        site = site_list[0]

    budget_pr_df = inverters_data.loc[:, inverters_data.columns.str.contains('Timestamp')]
    budget_pr_df[site + " Budget PR"] = [
        budget_pr.loc[site, str(datetime.strptime(str(row['Timestamp']), '%Y-%m-%d %H:%M:%S').date().replace(day=1))]
        for index, row in budget_pr_df.iterrows()]

    site_info = {}
    site_info['Site'] = site
    site_info['General Info'] = general_info.loc[site, :]
    site_info['Budget PR'] = budget_pr.loc[site, :]
    site_info['Budget PR table'] = budget_pr_df
    site_info['Component Info'] = component_data.loc[component_data['Site'] == site]

    # Get irradiance data and check for empty data
    irradiance_data = inverters_data.loc[:, inverters_data.columns.str.contains('Irradiance|Timestamp')]
    irradiance_data_nafilled = inverters_data.loc[:,
                               inverters_data.columns.str.contains('Irradiance|Timestamp')].fillna('No data')
    days_under_analysis = pd.to_datetime(irradiance_data['Timestamp']).dt.date.drop_duplicates()
    "months_under_analysis = pd.to_datetime(irradiance_data['Timestamp']).dt.month.drop_duplicates()"
    months_under_analysis = pd.to_datetime(irradiance_data['Timestamp']).apply(
        lambda x: x.strftime('%m-%Y')).drop_duplicates()

    site_info['Days'] = days_under_analysis
    site_info['Months'] = months_under_analysis

    # Active periods aka Irradiance > 20 W/m2
    active_irradiance = irradiance_data.loc[(irradiance_data[irradiance_curated_column] >= 20)]
    active_index = active_irradiance.index

    # To get periods without data
    empty_irradiance = irradiance_data_nafilled.loc[(irradiance_data_nafilled[irradiance_curated_column] == 'No data')]

    all_inverter_power_data_dict = {}

    for inverter in inverter_list:
        inverter_power_data_dict = {}
        capacity = float(site_info['Component Info'].loc[site_info['Component Info']['Component'] == inverter][
                             'Nominal Power DC'].values)

        inverter_power_data_dict['Power Data'] = inverters_data.loc[
            active_index, inverters_data.columns.str.contains(inverter + '|Timestamp')]

        # Get names of columns
        ac_power_column = inverter_power_data_dict['Power Data'].columns[
            inverter_power_data_dict['Power Data'].columns.str.contains('AC')].values[0]
        dc_power_column = inverter_power_data_dict['Power Data'].columns[
            inverter_power_data_dict['Power Data'].columns.str.contains('DC')].values[0]

        inverter_power_data_dict['Power Data']['Efficiency ' + inverter] = (
            inverter_power_data_dict['Power Data'][ac_power_column] /
            inverter_power_data_dict['Power Data'][dc_power_column])

        inverter_power_data_dict['Power Data']['Expected Power ' + inverter] = active_irradiance[
                                                                                   irradiance_curated_column].multiply(
            capacity / 1000) * budget_pr_df[site + " Budget PR"]

        inverter_power_data_dict['Power Data']['Ideal Power ' + inverter] = active_irradiance[
            irradiance_curated_column].multiply(capacity / 1000)

        inverter_power_data_dict['Power Data']['Irradiance'] = active_irradiance[irradiance_curated_column]

        all_inverter_power_data_dict[inverter] = inverter_power_data_dict
        power_data = inverter_power_data_dict['Power Data']

        # Create df with all data
        try:
            df_to_add = power_data.drop(columns=['Timestamp', 'Irradiance'])  # , 'Day', 'Month'
            all_inverter_data_df = pd.concat([all_inverter_data_df, df_to_add], axis=1)
        except NameError:
            print('Creating dataframe with all inverters')
            all_inverter_data_df = power_data

    return inverter_list, site_info, all_inverter_power_data_dict, \
           all_inverter_data_df, days_under_analysis, months_under_analysis


def remove_incidents_component_type(df, comp_type, column: str = "Related Component"):
    """ not done, add escape in case comp_type is not string """
    comp_type = str(comp_type)
    remove_index = []
    component_list = []
    for index, row in df.iterrows():
        rel_comp = df.loc[index, column]
        if comp_type in rel_comp:
            remove_index.append(index)
            component_list.append(rel_comp)

    df_final = df.drop(remove_index)
    df_final = df_final.reset_index(None, drop=True)

    return df_final


# </editor-fold>


# <editor-fold desc="Time related functions">

def get_percentage_of_timestamp(timestamp, rounded_timestamp, granularity: int = 15):
    difference = abs(rounded_timestamp - timestamp).seconds / 60

    percentage_of_timestamp = (granularity - difference) / granularity

    return percentage_of_timestamp


def read_analysis_df_and_correct_date(reportfiletemplate, date, roundto: int = 15):
    day = int(date[-2:])
    month = int(date[-5:-3])
    year = int(date[:4])

    freq = str(roundto) + 'min'
    df_incidents_analysis = pd.read_excel(reportfiletemplate, sheet_name='Analysis of CE', engine="openpyxl")
    df_tracker_analysis = pd.read_excel(reportfiletemplate, sheet_name='Analysis of tracker incidents',
                                        engine="openpyxl")

    timestamps = df_incidents_analysis['Time']
    timestamps = pd.Series(timestamps).dt.round(freq, 'shift_backward')
    df_incidents_analysis['Time'] = timestamps
    df_tracker_analysis['Time'] = timestamps

    for timestamp in timestamps:
        newtimestamp = timestamp.replace(year=year, month=month, day=day)
        df_incidents_analysis['Time'] = df_incidents_analysis['Time'].replace(timestamp, newtimestamp)
        df_tracker_analysis['Time'] = df_tracker_analysis['Time'].replace(timestamp, newtimestamp)

    # print(df_tracker_analysis)

    return df_incidents_analysis, df_tracker_analysis


def change_times_to_str(df, active: bool = False):
    df['Event Start Time'] = [str(time) for time in df['Event Start Time']]
    if not active:
        df['Event End Time'] = [str(time) for time in df['Event End Time']]

    return df


def remove_milliseconds(df, end_time: bool = False):
    """ Removes milliseconds from timestamps"""
    if end_time is True:
        df['Event Start Time'] = [str(timestamp) for timestamp in df['Event Start Time']]
        df['Event End Time'] = [str(timestamp) for timestamp in df['Event End Time']]

        for index, row in df.iterrows():
            stime = df.loc[index, 'Event Start Time']
            etime = df.loc[index, 'Event End Time']
            if "." in stime:
                dot_position = stime.index('.')
                df.loc[index, 'Event Start Time'] = str(stime[:dot_position])
            if "." in etime:
                dot_position = etime.index('.')
                df.loc[index, 'Event End Time'] = str(etime[:dot_position])
    else:
        df['Event Start Time'] = [str(timestamp) for timestamp in df['Event Start Time']]

        for index, row in df.iterrows():
            stime = df.loc[index, 'Event Start Time']
            if "." in stime:
                dot_position = stime.index('.')
                df.loc[index, 'Event Start Time'] = str(stime[:dot_position])

    return df


def remove_after_sunset_events(site_list, df_input, df_info_sunlight, active_df: bool = False, tracker: bool = False):
    if tracker is False:
        df_final = df_input
        # This script goes through the different site's dataframes one by one and checks all of the dataframe at once
        print(df_info_sunlight)
        for site in site_list:
            # get site dataframe
            df = df_input[site]
            # check if dataframe is not empty, in case it is then go straight to assigning the new df to the new df list

            if not df.empty:
                # Get index of site in info_sunlight dataframe
                index_site_array = df_info_sunlight[df_info_sunlight['Site'] == site].index.values
                index_site = int(index_site_array[0])

                # Get sunrise and sunset hour for a given site on the info_sunlight dataframe
                sunrise = df_info_sunlight.at[index_site, 'Time of operation start']
                if type(sunrise) == str:
                    sunrise = datetime.strptime(sunrise, '%Y-%m-%d %H:%M:%S')

                sunset = df_info_sunlight.at[index_site, 'Time of operation end']
                if type(sunset) == str:
                    sunset = datetime.strptime(sunset, '%Y-%m-%d %H:%M:%S')

                # Create test columns
                start_time_list = df['Event Start Time']
                start_time_list = [datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S') for timestamp in start_time_list]
                df['Event Start Time test'] = start_time_list
                if active_df is False:
                    end_time_list = df['Event End Time']
                    end_time_list = [datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S') for timestamp in end_time_list]
                    df['Event End Time test'] = end_time_list

                # Select all entries in dataframe that have start time before sunset or end time after sunrise a.k.a.
                # the wanted events;
                df_final_site = df.loc[df['Event Start Time test'] < sunset]
                if active_df is False:
                    df_final_site = df_final_site.loc[df['Event End Time test'] > sunrise]

                # get rid of the auxiliary columns
                df_final_site = df_final_site.drop(columns=['Event Start Time test'])
                if active_df is False:
                    df_final_site = df_final_site.drop(columns=['Event End Time test'])

                # add new dataframe to new dataframe list
                df_final[site] = df_final_site

    else:
        # this script goes through the incidents one by one to check which to keep because there are entries for
        # various sites in the same dataframe
        df = df_input
        if not df.empty:
            df['Outside of operation period'] = ""
            for index, row in df.iterrows():
                site_event = df.at[index, "Site Name"]
                index_site_array = df_info_sunlight[df_info_sunlight['Site'] == site_event].index.values
                index_site = int(index_site_array[0])

                sunrise = df_info_sunlight.at[index_site, 'Time of operation start']
                sunset = df_info_sunlight.at[index_site, 'Time of operation end']

                start_time_event = df.at[index, "Event Start Time"]
                if type(start_time_event) == str:
                    start_time_event = datetime.strptime(start_time_event, '%Y-%m-%d %H:%M:%S')
                if start_time_event >= sunset:
                    df.loc[index, 'Outside of operation period'] = "x"

                if active_df is False:
                    end_time_event = df.at[index, "Event End Time"]
                    if type(end_time_event) == str:
                        end_time_event = datetime.strptime(end_time_event, '%Y-%m-%d %H:%M:%S')
                    if end_time_event <= sunrise:
                        df.loc[index, 'Outside of operation period'] = "x"

                df_final = df.loc[df['Outside of operation period'] != "x"]
                df_final = df_final.drop(columns=['Outside of operation period'])
                df_final = df_final.reset_index(None, drop=True)
        else:
            df_final = df

    return df_final


def rounddatesactive_15m(site, df, freq: int = 15):
    """ Rounds date of active events (start time) to nearest 15m timestamp

    Returns df"""
    round_to = str(freq) + 'min'

    try:
        #to_round_startdate = pd.to_datetime(df['Event Start Time'], format='%Y-%m-%d %H:%M:%S')
        #to_round_startdate = df['Event Start Time']
        #to_round_startdate = [datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S') for timestamp in to_round_startdate]

        #rounded_startdate = pd.Series(to_round_startdate).dt.round(round_to, 'shift_backward')

        #rounded_startdate.index = df.index.to_list()

        df['Rounded Event Start Time'] = pd.to_datetime(df["Event Start Time"], format='%Y-%m-%d %H:%M:%S').dt.ceil("15min")

    except AttributeError:
        print('No new active events on this day for ' + site)

    return df


def rounddatesclosed_15m(site, df, freq: int = 15):
    """Rounds date of closed events (start time and end time) to nearest 15m timestamp

    Returns df """
    round_to = str(freq) + 'min'

    try:
        #to_round_startdate = pd.to_datetime(df['Event Start Time'], format='%Y-%m-%d %H:%M:%S')
        #to_round_startdate = df['Event Start Time']
        #to_round_startdate = [datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S') for timestamp in to_round_startdate]

        #to_round_enddate = pd.to_datetime(df['Event End Time'], format='%Y-%m-%d %H:%M:%S')

        #to_round_enddate = df['Event End Time']
        #to_round_enddate = [datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S') for timestamp in to_round_enddate]

        #rounded_startdate = pd.Series(to_round_startdate).dt.round(round_to, 'shift_forward')
        #rounded_enddate = pd.Series(to_round_enddate).dt.round(round_to, 'shift_forward')


        #rounded_startdate.index = df.index.to_list()
        #rounded_enddate.index = df.index.to_list()

        df['Rounded Event Start Time'] = pd.to_datetime(df["Event Start Time"], format='%Y-%m-%d %H:%M:%S').dt.ceil("15min")
        df['Rounded Event End Time'] = pd.to_datetime(df["Event End Time"], format='%Y-%m-%d %H:%M:%S').dt.ceil("15min")

    except AttributeError:
        print('No new closed events on this day for ' + site)

    return df


def verify_read_time_of_operation(site, day, stime, etime):
    sg.theme('DarkAmber')  # Add a touch of color
    # All the stuff inside your window.
    layout = [[sg.Text('Are these hours correct on the ' + str(day) + ' on ' + site + '?')],
              [sg.HorizontalSeparator(pad=((10, 10), (2, 10)))],
              [sg.Text('Sunrise hour: '), sg.Text(str(stime))],
              [sg.Text('Sunset hour: '), sg.Text(str(etime))],
              [sg.Button('Submit'), sg.Button('Change hours'), sg.Exit()]]

    # Create the Window
    window = sg.Window('Daily Monitoring Report', layout)
    # Event Loop to process "events" and get the "values" of the inputs

    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == 'Exit':  # if user closes window or clicks exit
            break
        if event == 'Submit':
            print('Submitted for ' + site + ':')
            print(stime)
            print(etime)
            window.close()

            return stime, etime

        if event == 'Change hours':
            stime, etime = inputs.input_time_operation_site(site, str(day))
            window.close()
            return stime, etime

    window.close()

    return stime, etime


# </editor-fold>


# <editor-fold desc="Dataframe creation">

def filter_site_selection(df, site_selection):
    """From a dataframe of a report containing all incidents from various sites, this script filters it out for
    the given site_selection

    USAGE: create_dfs(original_df,site_selection)

    Returns df_corrected"""

    sites = df['Site Name']
    corrected_sites = [correct_site_name(site) for site in sites]
    df['Site Name'] = corrected_sites

    df_corrected = df[df["Site Name"].isin(site_selection)]

    return df_corrected


def create_dfs(df, site_selection, min_dur: int = 15, roundto: int = 15):
    """From a dataframe of a report containing all incidents from various sites, this script creates a dictionary
    (list with keys and values (ex: 'LSBP - Grants' : df_active_events_Grants) for ACTIVE and CLOSED EVENTS containing
    all sites present in the original dataframe. Independently of the number of sites in the input dataframe

    USAGE: create_dfs(original_df,minimum duration of events (default = 15), auxiliary_df_for_block_identification)

    Returns site_list,df_list_active, df_list_closed"""

    df_closed_all = filter_notprod_and_duration(df,
                                                min_dur)  # creates dataframe with closed, not producing incidents with a minimum specified duration
    df_closed_all = remove_milliseconds(df_closed_all, end_time=True)  # removes milliseconds
    # append_df_to_excel('test.xlsx', df_closed_all, sheet_name='test')

    df_active_all = create_active_events_df(df)  # creates dataframe with closed, not producing incidents
    df_active_all = remove_milliseconds(df_active_all)  # removes milliseconds

    df_closed_all = remove_incidents_component_type(df_closed_all, 'Feeder')  # removes incidents of inverter modules
    df_active_all = remove_incidents_component_type(df_active_all, 'Feeder')

    # Add capacity of each component
    # df_list_active, df_list_closed = create_df_list(site_selection)
    df_list_active = {}
    df_list_closed = {}

    for site in site_selection:
        # Create active df for a given site
        df_active = df_active_all.loc[df_active_all['Site Name'] == site]
        df_active = df_active.reset_index(None, drop=True)
        df_active = rounddatesactive_15m(site, df_active, freq=roundto)

        df_list_active[site] = df_active

        # Create closed df for a given site
        df_closed = df_closed_all.loc[df_closed_all['Site Name'] == site]
        df_closed = df_closed.reset_index(None, drop=True)
        df_closed = rounddatesclosed_15m(site, df_closed, freq=roundto)

        df_list_closed[site] = df_closed

    return df_list_active, df_list_closed


def create_tracker_dfs(df_all, site_selection, df_general_info_calc, roundto: int = 15):
    df_tracker_closed = closedtrackerdf(df_all, df_general_info_calc)
    df_tracker_active = activetrackerdf(df_all, df_general_info_calc)

    df_tracker_active = remove_milliseconds(df_tracker_active)
    df_tracker_closed = remove_milliseconds(df_tracker_closed, end_time=True)

    df_tracker_closed = remove_incidents_component_type(df_tracker_closed, 'TrackerModeEnabled', 'State')
    df_tracker_active = remove_incidents_component_type(df_tracker_active, 'TrackerModeEnabled', 'State')

    df_tracker_list_active = {}
    df_tracker_list_closed = {}

    for site in site_selection:
        # Create active df for a given site
        df_tracker_active_site = df_tracker_active.loc[df_tracker_active['Site Name'] == site]
        df_tracker_active_site = df_tracker_active_site.reset_index(None, drop=True)
        df_tracker_active_site = rounddatesactive_15m(site, df_tracker_active_site, freq=roundto)

        df_tracker_list_active[site] = df_tracker_active_site

        # Create closed df for a given site
        df_tracker_closed_site = df_tracker_closed.loc[df_tracker_closed['Site Name'] == site]
        df_tracker_closed_site = df_tracker_closed_site.reset_index(None, drop=True)
        df_tracker_closed_site = rounddatesclosed_15m(site, df_tracker_closed_site, freq=roundto)

        df_tracker_list_closed[site] = df_tracker_closed_site

    return df_tracker_list_active, df_tracker_list_closed


def activetrackerdf(df_15m, df_tracker_info_calc):
    """This script  creates a dataframe containing all active tracker incidents from a given report.
    It doesn't separate the data by sites. All data is gathered in the same dataframe

    Returns df_tracker_active"""

    df_15m_tracker = df_15m.loc[(df_15m["Related Component"].str.contains('Tracker|TRACKER')) |
                                (df_15m["State"] == 'Tracker target availability')]
    df_15m_tracker["Tracker"] = ['Yes'] * len(df_15m_tracker['Related Component'])

    df_15m_tracker['Capacity Related Component'] = [
        df_tracker_info_calc.loc[site, 'Avg. capacity per tracker (kW)']
        for site in df_15m_tracker['Site Name']]

    df_tracker_active = df_15m_tracker.loc[pd.isna(df_15m_tracker['Event End Time'])]
    df_tracker_active = df_tracker_active.reset_index(None, drop=True)

    return df_tracker_active


def closedtrackerdf(df_15m, df_tracker_info_calc):
    """This script  creates a dataframe containing all active tracker incidents from a given report.
    It doesn't separate the data by sites. All data is gathered in the same dataframe
    This script doesn't filter the incidents by duration.

    Returns df_15m_tracker_active"""

    df_15m_tracker = df_15m.loc[(df_15m["Related Component"].str.contains('Tracker|TRACKER')) |
                                (df_15m["State"] == 'Tracker target availability')]
    df_15m_tracker["Tracker"] = ['Yes'] * len(df_15m_tracker['Related Component'])

    df_15m_tracker['Capacity Related Component'] = [df_tracker_info_calc.loc[site, 'Avg. capacity per tracker (kW)']
                                                    for site in df_15m_tracker['Site Name']]

    df_15m_tracker_closed = df_15m_tracker.loc[df_15m_tracker['Event End Time'].notnull()]

    df_15m_tracker_dur15m = df_15m_tracker_closed.loc[
        df_15m_tracker_closed['Duration (h)'] > 0.249]  # filter by duration >15 minute
    df_15m_tracker_final = df_15m_tracker_dur15m.reset_index(None, drop=True)  # reset index

    return df_15m_tracker_final


def create_df_list(site_selection):
    """This script creates two dictionaries and a list.
     The dictionaries contain all sites from a given report and their active and closed events dataframes
     The list contains all sites present in the report

     Returns site_list, df_list_active, df_list_closed"""

    """for site in site_selection:
        index_site = site_selection.index(site)
        site = correct_site_name(site)
        site_selection[index_site] = site"""

    for site in site_selection:
        if "LSBP - " in site or "LSBP – " in site:
            onlysite = site[7:]
        else:
            onlysite = site
        onlysite = onlysite.replace(" ", "")
        df_name_active = "df_15m_" + onlysite + "_active"
        df_name_closed = "df_15m_" + onlysite + "_closed"

        try:
            if site not in df_list_active.keys():
                df_list_active[site] = df_name_active

        except NameError:
            df_list_active = {site: df_name_active}

        try:
            if site not in df_list_closed.keys():
                df_list_closed[site] = df_name_closed

        except NameError:
            df_list_closed = {site: df_name_closed}

    return df_list_active, df_list_closed


def create_active_events_df(df):
    """This script creates a dataframe containing all active event from a given report
    It doesn't filter the incidents by site, gathering them all in a single dataframe

    Returns df_active"""

    df_15m_notprod = df.loc[df['Component Status'] == 'Not Producing']
    df_active = df_15m_notprod.loc[df_15m_notprod['Event End Time'].isnull()]

    return df_active


# </editor-fold>


# <editor-fold desc="Dataframe completion">

def complete_dataset_capacity_data(df_list, all_component_data):
    for site in df_list.keys():
        incidents_site = df_list[site]
        # print(type(incidents_site))
        if not type(incidents_site) == str:
            for index, row in incidents_site.iterrows():
                site = row['Site Name']
                component = row['Related Component']

                try:
                    capacity = all_component_data.loc[(all_component_data['Site'] == site)
                                                      & (all_component_data['Component'] == component)][
                        "Nominal Power DC"].values[0]
                except IndexError:
                    capacity = "NA"

                # Add capacity
                incidents_site.loc[index, 'Capacity Related Component'] = capacity

            df_list[site] = incidents_site

    return df_list


def complete_dataset_existing_incidents(df_list, df_dmr):
    for site in df_list.keys():
        print("Completing dataset on " + site)
        incidents_site = df_list[site]
        df_raw_columns = incidents_site.columns.to_list()

        df_dmr_site = df_dmr.loc[df_dmr['Site Name'] == site]

        if type(df_dmr_site) == str:
            print("No previous active events")

        elif type(incidents_site) == str:
            print("No active events, adding previously active events")
            incidents_site = pd.concat([incidents_site, df_dmr_site])[df_raw_columns]

        else:
            incidents_site = pd.concat([incidents_site, df_dmr_site])[df_raw_columns]

        df_list[site] = incidents_site

    return df_list



def fill_events_analysis_dataframe(df_analysis, df_info_sunlight):
    max_percentage = "{:.2%}".format(1)
    site_list = df_info_sunlight['Site']
    for site in site_list:
        index_site_array = df_info_sunlight[df_info_sunlight['Site'] == site].index.values
        index_site = int(index_site_array[0])
        stime = df_info_sunlight.loc[index_site, 'Time of operation start']
        etime = df_info_sunlight.loc[index_site, 'Time of operation end']

        index_mint = df_analysis[
            df_analysis['Time'] == stime].index.values  # gets starting time row index
        int_index_mint = index_mint[0]  # turns index from numpy.ndarray to integer

        index_maxt = df_analysis[
            df_analysis['Time'] == etime].index.values  # gets ending time row index
        int_index_maxt = index_maxt[0]

        for index in range(int_index_mint, int_index_maxt):
            df_analysis.loc[index, site] = max_percentage

    return df_analysis


def comprehensive_description(df):
    a = 1
    return a


def describe_incidents(df, df_info_sunlight, active_events: bool = False, tracker: bool = False):
    site_list = df.keys()
    for site in site_list:
        df_events = df[site]
        index_site_array = df_info_sunlight[df_info_sunlight['Site'] == site].index.values
        index_site = int(index_site_array[0])
        sunrise_time = df_info_sunlight.loc[index_site, 'Time of operation start']
        site_capacity = df_info_sunlight.loc[index_site, 'Capacity']

        if active_events is False:
            print('Describing closed incidents of ' + site)
            for index, row in df_events.iterrows():
                rel_comp = df_events.at[index, 'Related Component']
                cap_rel_comp = df_events.at[index, 'Capacity Related Component']
                if tracker:
                    status = "off position"
                else:
                    status = df_events.at[index, 'Component Status']

                duration = df_events.at[index, 'Duration (h)']

                start_date = df_events.at[index, 'Rounded Event Start Time']
                end_date = df_events.at[index, 'Rounded Event End Time']

                start_date_short = start_date.strftime("%b-%d")
                end_date_short = end_date.strftime("%b-%d")

                event_time_hour = end_date.hour
                event_time_minute = end_date.minute
                if event_time_minute == 0:
                    event_time = str(event_time_hour) + ':0' + str(event_time_minute)
                else:
                    event_time = str(event_time_hour) + ':' + str(event_time_minute)

                if start_date == sunrise_time and duration < 2:
                    description = "• " + str(rel_comp) + ' started late at ~' + str(event_time) + ' (closed)'

                elif start_date.day != end_date.day:
                    description = "• " + str(rel_comp) + ' was ' + status.lower() + ' from' + start_date_short + \
                                  'until ' + end_date_short + ' at ~' + str(event_time) + ' (' + \
                                  "{:.2%}".format(cap_rel_comp/site_capacity) + 'of site capacity affected)'
                else:
                    description = "• " + str(rel_comp) + ' was ' + status.lower() + ' on the ' + str(start_date.day)\
                                  + ' for ~' + str(duration) + ' hours (' + \
                                  "{:.2%}".format(cap_rel_comp/site_capacity) + 'of site capacity affected)'

                df_events.loc[index, 'Comments'] = description

            df[site] = df_events

        else:
            print('Describing active incidents of ' + site)
            for index, row in df_events.iterrows():
                rel_comp = df_events.at[index, 'Related Component']
                start_date = df_events.at[index, 'Rounded Event Start Time']
                day = start_date.day
                month = start_date.month
                date = str(day) + '/' + str(month)
                if tracker:
                    status = "off position"
                else:
                    status = df_events.at[index, 'Component Status']

                description = "• " + str(rel_comp) + ' is ' + status.lower() + ' (open since ' + \
                              start_date.strftime("%b-%d") + ')'

                df_events.loc[index, 'Comments'] = description

            df[site] = df_events


    return df


# </editor-fold>


# <editor-fold desc="ET Functions">


def create_fmeca_dataframes_for_validation(fmeca_data):
    """From FMECA Table creates all dataframes needed for data_validation.
    Structures Faults, Fault Component, Failure Mode, Failure Mechanism, Category and Subcategory"""

    # data validation for FMECA

    # next level is dependent on combination of previous levels
    faults_fmeca = list(set(fmeca_data['Fault'].to_list()))
    fault_component_fmeca = dict(
        (fault, list(set(fmeca_data.loc[fmeca_data['Fault'] == fault]['Fault Component'].to_list()))) for fault in
        faults_fmeca)
    failure_mode_fmeca = dict(((fault, fault_comp), list(set(
        fmeca_data.loc[(fmeca_data['Fault'] == fault) & (fmeca_data['Fault Component'] == fault_comp)][
            'Failure Mode'].to_list()))) for fault, fault_comps in fault_component_fmeca.items() for fault_comp in
                              fault_comps)

    failure_mechanism_fmeca = \
        dict((fault_and_comp + (fail_mode,), list(set(fmeca_data.loc[(fmeca_data['Fault'] == fault_and_comp[0]) &
                                                                     (fmeca_data['Fault Component'] ==
                                                                      fault_and_comp[1])
                                                                     & (fmeca_data['Failure Mode']
                                                                        == fail_mode)]['Failure Mechanism'].to_list())))
             for fault_and_comp, fail_modes in failure_mode_fmeca.items() for fail_mode in fail_modes)

    category_fmeca = \
        dict((fault_and_comp_mode + (fail_mec,),
              list(set(fmeca_data.loc[(fmeca_data['Fault'] == fault_and_comp_mode[0]) &
                                      (fmeca_data['Fault Component'] == fault_and_comp_mode[1]) &
                                      (fmeca_data['Failure Mode'] == fault_and_comp_mode[2]) &
                                      (fmeca_data['Failure Mechanism'] == fail_mec)]['Category'].to_list())))
             for fault_and_comp_mode, fail_mecs in failure_mechanism_fmeca.items() for fail_mec in fail_mecs)

    subcategory_fmeca = \
        dict((fault_and_comp_mode_mec + (cat,),
              list(set(fmeca_data.loc[(fmeca_data['Fault'] == fault_and_comp_mode_mec[0]) &
                                      (fmeca_data['Fault Component'] == fault_and_comp_mode_mec[1]) &
                                      (fmeca_data['Failure Mode'] == fault_and_comp_mode_mec[2]) &
                                      (fmeca_data['Failure Mechanism'] == fault_and_comp_mode_mec[3]) &
                                      (fmeca_data['Category'] == cat)]['Subcategory'].to_list())))
             for fault_and_comp_mode_mec, cats in category_fmeca.items() for cat in cats)

    # Change multi-level options' keys to have all dependencies on key
    fault_component_fmeca_newkeys = dict(
        (key, key.replace(" ", "_").replace("-", "_")) for key in fault_component_fmeca.keys())
    failure_mode_fmeca_newkeys = dict(
        (key, "_".join(key).replace(" ", "_").replace("-", "_")) for key in failure_mode_fmeca.keys())
    failure_mechanism_fmeca_newkeys = dict(
        (key, "_".join(key).replace(" ", "_").replace("-", "_")) for key in failure_mechanism_fmeca.keys())
    category_fmeca_newkeys = dict(
        (key, "_".join(key).replace(" ", "_").replace("-", "_")) for key in category_fmeca.keys())
    subcategory_fmeca_newkeys = dict(
        (key, "_".join(key).replace(" ", "_").replace("-", "_")) for key in subcategory_fmeca.keys())

    fault_component_fmeca = rename_dict_keys(fault_component_fmeca,
                                             fault_component_fmeca_newkeys)

    failure_mode_fmeca = rename_dict_keys(failure_mode_fmeca, failure_mode_fmeca_newkeys)

    failure_mechanism_fmeca = rename_dict_keys(failure_mechanism_fmeca,
                                               failure_mechanism_fmeca_newkeys)

    category_fmeca = rename_dict_keys(category_fmeca, category_fmeca_newkeys)

    subcategory_fmeca = rename_dict_keys(subcategory_fmeca, subcategory_fmeca_newkeys)

    # Create dfs

    df_faults_fmeca = pd.DataFrame(data={'Faults': faults_fmeca})
    df_fault_component_fmeca = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in fault_component_fmeca.items()]))
    df_failure_mode_fmeca = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in failure_mode_fmeca.items()]))
    df_failure_mechanism_fmeca = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in failure_mechanism_fmeca.items()]))
    df_category_fmeca = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in category_fmeca.items()]))
    df_subcategory_fmeca = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in subcategory_fmeca.items()]))

    dict_fmeca_shapes = {'Faults': (df_faults_fmeca, df_faults_fmeca.shape),
                         'Failure_Component': (df_fault_component_fmeca, df_fault_component_fmeca.shape),
                         'Failure_Mode': (df_failure_mode_fmeca, df_failure_mode_fmeca.shape),
                         'Failure_Mechanism': (df_failure_mechanism_fmeca, df_failure_mechanism_fmeca.shape),
                         'Category': (df_category_fmeca, df_category_fmeca.shape),
                         'Subcategory': (df_subcategory_fmeca, df_subcategory_fmeca.shape)}

    return dict_fmeca_shapes


def match_df_to_event_tracker(df, component_data, fmeca_data, active: bool = False,
                              simple_match: bool = False, tracker: bool = False):
    desired_columns_components = ["ID", "Site Name", "Related Component", "Capacity Related Component",
                                  "Component Status", "Event Start Time", "Event End Time", "Duration (h)",
                                  "Active Hours (h)", "Energy Lost (MWh)", "Comments","Remediation", "Fault",
                                  "Fault Component", "Failure Mode", "Failure Mechanism", "Category", "Subcategory",
                                  "Resolution Category", "Excludable", "Exclusion Start Time", "Exclusion End Time",
                                  "Excludable Category", "Exclusion Rationale", "Incident Status",
                                  "Categorization Status"]
    if not df.empty:
        if simple_match is False:
            curtailment_fmeca = fmeca_data.loc[fmeca_data['Failure Mode'] == 'Curtailment']

            df['Site Name'] = [correct_site_name(name) for name in df['Site Name']]
            df['Related Component'] = [correct_site_name(name)
                                       for name in df['Related Component']]

            # print(df['Site Name'],df['Related Component'])

            # Add ID column
            if "ID" not in df.columns:
                #df = change_times_to_str(df, active=active)

                # test if all entries to add have an entry on the general info file
                tuple_list_toadd = list(
                    set([(row['Site Name'], row['Related Component']) for index, row in df.iterrows()]))
                tuple_list_componentdata = [(row['Site'], row['Component']) for index, row in component_data.iterrows()]

                does_not_exist = [x for x in tuple_list_toadd if x not in tuple_list_componentdata]

                if not does_not_exist:
                    df.insert(0, "ID", [component_data.loc
                                        [(component_data['Site'] == df.loc[index, 'Site Name']) &
                                         (component_data['Component'] == df.loc[index, 'Related Component'])]['ID'].values[0]
                                        + '-' + str(df.loc[index, 'Event Start Time']).replace(" ", "T").replace("-", "").replace(":", "")
                                        for index, row in df.iterrows()])

                else:
                    print(does_not_exist)
                    print(df.loc[df['Related Component'] == does_not_exist[1]])
                    """print(tuple_list_toadd)
                    print(tuple_list_componentdata)"""
                    sg.popup("These components do not exist in the general info file: " + str(does_not_exist))
                    sys.exit("These components do not exist in the general info file")

        # Add rest of the columns
        for column in desired_columns_components:
            if column not in df.columns:
                df[column] = ""
            elif column == 'Incident Status':
                if active is True:
                    df[column] = "Open"
                else:
                    df[column] = "Closed"
            elif column == 'Categorization Status':
                if active is True:
                    df[column] = "Pending"
                else:
                    df[column] = ["Pending" if status == "" else status for status in df[column]]

        df.drop_duplicates(subset=['ID'], inplace=True, ignore_index=True)  # .reset_index(drop=True, inplace=True)

        if simple_match is False and tracker is False:
            try:
                for index, row in df.loc[df['Curtailment Event'] == 'x'].iterrows():
                    df.loc[index, 'Fault'] = curtailment_fmeca['Fault'].values[0]
                    df.loc[index, 'Fault Component'] = curtailment_fmeca['Fault Component'].values[0]
                    df.loc[index, 'Failure Mode'] = curtailment_fmeca['Failure Mode'].values[0]
                    df.loc[index, 'Failure Mechanism'] = curtailment_fmeca['Failure Mechanism'].values[0]
                    df.loc[index, 'Category'] = curtailment_fmeca['Category'].values[0]
                    df.loc[index, 'Subcategory'] = curtailment_fmeca['Subcategory'].values[0]
                    df.loc[index, 'Resolution Category'] = "Reset"
                    df.loc[index, 'Excludable'] = "Yes"
                    df.loc[index, 'Excludable Category'] = "Curtailment"
                    df.loc[index, 'Exclusion Rationale'] = "Curtailment"
                    df.loc[index, 'Incident Status'] = "Closed"
                    df.loc[index, 'Categorization Status'] = "Completed"

            except KeyError:
                pass
            # print(df['Related Component'])
            for index, row in df.loc[(df['Related Component'].str.contains('CB|DC|String'))].iterrows():
                df.loc[index, 'Excludable'] = "Yes"
                df.loc[index, 'Excludable Category'] = "Sub-Inverter Level"
                df.loc[index, 'Exclusion Rationale'] = "Sub-Inverter Level"

        df.drop(columns=df.columns.difference(desired_columns_components), inplace=True)

        df_final = df[desired_columns_components]

    else:
        df_final = pd.DataFrame(columns=desired_columns_components)

    return df_final

# </editor-fold>

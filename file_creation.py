import pandas as pd
from append_df_to_excel import append_df_to_excel
from datetime import datetime
import datetime as dt
import os
import re
import sys
import PySimpleGUI as sg
import openpyxl
import xlsxwriter
import perfonitor.calculations as calculations
import perfonitor.data_acquisition as data_acquisition
import perfonitor.inputs as inputs
import perfonitor.data_treatment as data_treatment
import perfonitor.data_analysis as data_analysis


# File creation/edit/removal-------------------------------------------------------------------------------------

def add_incidents_to_excel(dest_file, site_list, df_list_active,
                           df_list_closed, df_info_sunlight, final_irradiance_data):
    """USAGE: add_incidents_to_excel(destiny_file,site_list,df_list_active,df_list_closed)"""

    append_df_to_excel(dest_file, df_info_sunlight, sheet_name='Info', startrow=0)
    append_df_to_excel(dest_file, final_irradiance_data, sheet_name='Irradiance', startrow=0)

    for site in site_list:
        if "LSBP - " in site or "LSBP – " in site:
            onlysite = site[7:]
        else:
            onlysite = site
        if onlysite[-1:] == " ":
            active_sheet_name = onlysite + 'Active'
            closed_sheet_name = onlysite[:len(onlysite) - 1]
        else:
            active_sheet_name = onlysite + ' Active'
            closed_sheet_name = onlysite

        df_active = df_list_active[site]
        df_closed = df_list_closed[site]

        df_closed['Status of incident'] = 'Closed'
        df_active['Status of incident'] = 'Active'
        df_active['Action required'] = ''

        append_df_to_excel(dest_file, df_closed, sheet_name=closed_sheet_name)

        print('Active events of ' + site + ' added')
        append_df_to_excel(dest_file, df_active, sheet_name=active_sheet_name)
        print('Closed events of ' + site + ' added')

    return


def add_tracker_incidents_to_excel(dest_tracker_file, df_tracker_active, df_tracker_closed, df_tracker_info):
    """USAGE: add_tracker_incidents_to_excel(dest_file, df_tracker_active, df_tracker_closed, df_tracker_info)"""

    append_df_to_excel(dest_tracker_file, df_tracker_info, sheet_name='Trackers info', startrow=0)
    print('Tracker Info added')

    df_tracker_closed['Status of incident'] = 'Closed'
    df_tracker_active['Status of incident'] = 'Active'
    df_tracker_active['Action required'] = ''

    append_df_to_excel(dest_tracker_file, df_tracker_active, sheet_name='Active tracker incidents')
    print('Active tracker incidents added')
    append_df_to_excel(dest_tracker_file, df_tracker_closed, sheet_name='Closed tracker incidents')
    print('Closed tracker incidents added')

    return


def add_events_to_final_report(reportfile_path, df_list_active, df_list_closed, df_tracker_active, df_tracker_closed):
    final_active_events_list = pd.concat(list(df_list_active.values()))
    final_closed_events_list = pd.concat(list(df_list_closed.values()))

    if not final_active_events_list.empty:
        append_df_to_excel(reportfile_path, final_active_events_list, sheet_name='Active Events', startrow=0)
        print('Active events added')
    else:
        print('No active events to be added')

    if not final_closed_events_list.empty:
        append_df_to_excel(reportfile_path, final_closed_events_list, sheet_name='Closed Events', startrow=0)
        print('Closed events added')
    else:
        print('No closed events to be added')

    if not df_tracker_active.empty:
        append_df_to_excel(reportfile_path, df_tracker_active, sheet_name='Active tracker incidents', startrow=0)
        print('Tracker active events added')
    else:
        print('No tracker active events to be added')

    if not df_tracker_closed.empty:
        append_df_to_excel(reportfile_path, df_tracker_closed, sheet_name='Closed tracker incidents', startrow=0)
        print('Tracker closed events added')
    else:
        print('No tracker closed events to be added')

    return


def add_analysis_to_reportfile(reportfile, df_incidents_analysis, df_tracker_analysis, df_info_sunlight):
    # Add Info sheet
    append_df_to_excel(reportfile, df_info_sunlight, sheet_name='Info', startrow=0)

    # Add component failure analysis
    append_df_to_excel(reportfile, df_incidents_analysis, sheet_name='Analysis of CE', startrow=0)

    # Add component failure analysis
    append_df_to_excel(reportfile, df_tracker_analysis, sheet_name='Analysis of tracker incidents', startrow=0)

    return


def update_dump_file(irradiance_files, all_irradiance_file, data_type: str = 'Irradiance'):
    df_all_irradiance = pd.read_excel(all_irradiance_file, engine='openpyxl')

    df_irradiance_day_list = [pd.read_excel(file, engine='openpyxl') for file in irradiance_files]
    df_all_irradiance_list = df_irradiance_day_list.append(df_all_irradiance)

    df_all_irradiance_new = pd.concat(df_irradiance_day_list)
    df_all_irradiance_new['Timestamp'] = [datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S') for timestamp in
                                          df_all_irradiance_new['Timestamp']]

    df_all_irradiance_new = df_all_irradiance_new.loc[:, ~df_all_irradiance_new.columns.str.contains('^Unnamed')]. \
        drop_duplicates(subset=['Timestamp'], keep='first', ignore_index=True).sort_values(
        by=['Timestamp'], ascending=[True], ignore_index=True)

    writer_irr = pd.ExcelWriter(all_irradiance_file, engine='xlsxwriter',
                                engine_kwargs={'options': {'strings_to_numbers': True}})

    workbook_irr = writer_irr.book

    df_all_irradiance_new.to_excel(writer_irr, sheet_name='All ' + str(data_type), index=False)

    writer_irr.save()

    return df_all_irradiance_new


def dmr_create_incidents_files(alarm_report_path, irradiance_file_path, geography, date, site_selection):
    geography_folder = os.path.dirname(alarm_report_path)
    print(site_selection)
    report_files_dict = {}
    # print("this is dir: " + dir)

    previous_date = datetime.strptime(date, '%Y-%m-%d') - dt.timedelta(days=1)
    prev_day = str(previous_date.day) if previous_date.day >= 10 else str(0) + str(previous_date.day)
    prev_month = str(previous_date.month) if previous_date.month >= 10 else str(0) + str(previous_date.month)

    report_template_path = geography_folder + '/Info&Templates/Reporting_' + geography + '_Sites_Template.xlsx'
    general_info_path = geography_folder + '/Info&Templates/General Info ' + geography + '.xlsx'
    event_tracker_path = geography_folder + '/Event Tracker/Event Tracker ' + geography + '.xlsx'
    # previous_dmr_path = geography_folder + '/Reporting_' + geography + '_Sites_' + str(previous_date.date()).replace("-", "") + '.xlsx'

    folder_content = os.listdir(geography_folder)
    previous_dmr_path = 'Reporting_' + geography + '_Sites_' + str(previous_date.date()).replace("-", "")
    report_file_list = [geography_folder + '/' + file for file in folder_content if previous_dmr_path in file]

    print(report_file_list)

    # READ FILES AND EXTRACT RAW DATAFRAMES
    print('Reading Daily Alarm Report...')
    df_all, incidents_file, tracker_incidents_file, irradiance_df, prev_active_events, prev_active_tracker_events \
        = data_acquisition.read_daily_alarm_report(alarm_report_path, irradiance_file_path, event_tracker_path,
                                                   report_file_list)
    print('Daily Alarm Report read!')
    print('newfile: ' + incidents_file)
    print('newtrackerfile: ' + tracker_incidents_file)
    print('Reading trackers info...')
    df_general_info, df_general_info_calc, all_component_data = data_acquisition.read_general_info(general_info_path)
    print('Trackers info read!')

    # DIVIDE RAW DATAFRAMES INTO LIST OF DATAFRAMES BY SITE
    print('Filtering incidents list for site selection...')
    df_all = data_treatment.filter_site_selection(df_all, site_selection)
    print('Creating incidents dataframes list...')
    df_list_active, df_list_closed = data_treatment.create_dfs(df_all, site_selection, min_dur=1, roundto=1)
    print('Incidents dataframes list created')
    print('Creating tracker dataframes...')
    df_tracker_active, df_tracker_closed = data_treatment.create_tracker_dfs(df_all, df_general_info_calc, roundto=1)
    print('Tracker dataframes created')
    print('Please set time of operation')
    df_info_sunlight, final_irradiance_data = data_acquisition.read_time_of_operation_new(irradiance_df, site_selection,
                                                                                          df_general_info,
                                                                                          withmean=False)

    # df_info_sunlight = msp.set_time_of_operation(Report_template_path, site_list, date)
    print('Removing incidents occurring after sunset')
    df_list_closed = data_treatment.remove_after_sunset_events(site_selection, df_list_closed, df_info_sunlight)
    df_list_active = data_treatment.remove_after_sunset_events(site_selection, df_list_active,
                                                               df_info_sunlight, active_df=True)

    df_tracker_closed = data_treatment.remove_after_sunset_events(site_selection, df_tracker_closed,
                                                                  df_info_sunlight, tracker=True)

    df_tracker_active = data_treatment.remove_after_sunset_events(site_selection, df_tracker_active,
                                                                  df_info_sunlight, active_df=True, tracker=True)
    print('Adding component capacities')
    # ADD CAPACITIES TO DFS
    df_list_closed = data_treatment.complete_dataset_capacity_data(df_list_closed, all_component_data)
    df_list_active = data_treatment.complete_dataset_capacity_data(df_list_active, all_component_data)

    # JOIN INCIDENT TABLES AND DMR TABLES
    df_list_active = data_treatment.complete_dataset_existing_incidents(df_list_active, prev_active_events)
    df_tracker_active = pd.concat([df_tracker_active, prev_active_tracker_events])[df_tracker_active.columns.to_list()]

    # CREATE INCIDENTS FILE

    print('Creating Incidents file...')
    print(incidents_file)
    add_incidents_to_excel(incidents_file, site_selection, df_list_active, df_list_closed, df_info_sunlight,
                           final_irradiance_data)
    print('Incidents file created!')
    print('Creating tracker incidents file...')
    add_tracker_incidents_to_excel(tracker_incidents_file, df_tracker_active, df_tracker_closed, df_general_info)
    print('Tracker incidents file created!')

    return incidents_file, tracker_incidents_file, all_component_data


def dmrprocess1(site_selection: list = []):
    sg.theme('DarkAmber')  # Add a touch of color
    # All the stuff inside your window.
    layout = [[sg.Text('Enter date of report you want to analyse', pad=((2, 10), (2, 5)))],
              [sg.CalendarButton('Choose date', target='-CAL-', format="%Y-%m-%d"),
               sg.In(key='-CAL-', text_color='black', size=(16, 1), enable_events=True, readonly=True, visible=True)],
              [sg.Text('Choose Alarm report', pad=((0, 10), (10, 2)))],
              [sg.FileBrowse(target='-FILE-'),
               sg.In(key='-FILE-', text_color='black', size=(20, 1), enable_events=True, readonly=True, visible=True)],
              [sg.Text('Choose Irradiance file', pad=((0, 10), (10, 2)))],
              [sg.FileBrowse(target='-IRRFILE-'),
               sg.In(key='-IRRFILE-', text_color='black', size=(20, 1), enable_events=True, readonly=True,
                     visible=True)],
              [sg.Text('Enter geography ', pad=((0, 10), (10, 2)))],
              [sg.Combo(['AUS', 'ES', 'USA'], size=(4, 3), readonly=True, key='-GEO-', pad=((5, 10), (2, 10)))],
              [sg.Button('Create Incidents List'), sg.Exit()]]

    # Create the Window
    window = sg.Window('Daily Monitoring Report', layout, modal=True)
    # Event Loop to process "events" and get the "values" of the inputs
    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == 'Exit':  # if user closes window or clicks exit
            window.close()
            return "No File", "No File", ["No site list"], "PT", "27-03-1996", "No Data"

        if event == 'Create Incidents List':
            date = values['-CAL-']  # date is string
            Alarm_report_path = values['-FILE-']
            irradiance_file_path = values['-IRRFILE-']

            report_name = os.path.basename(Alarm_report_path)
            geography_report_match = re.search(r'\w+?_', report_name)
            geography_report = geography_report_match.group()[:-1]
            geography = values['-GEO-']

            # print(date[:4])
            # print(date[5:7])
            # print(date[-2:])
            print(Alarm_report_path)
            print(geography)
            print(geography_report)
            print(site_selection)
            if "Daily" and "Alarm" and "Report" in Alarm_report_path and geography == geography_report and \
                "Irradiance" in irradiance_file_path:

                incidents_file, tracker_incidents_file, all_component_data = dmr_create_incidents_files(
                    Alarm_report_path, irradiance_file_path, geography, date, site_selection)
                sg.popup('All incident files are ready for approval', no_titlebar=True)
                window.close()
                return incidents_file, tracker_incidents_file, geography, date, all_component_data

            elif not geography == geography_report:
                msg = 'Selected Geography ' + geography + ' does not match geography from report ' + geography_report
                sg.popup(msg, title="Error with the selections")
                # print('Selected Geography ' + geography + ' does not match geography from report ' + geography_report)
            elif not "Daily" and "Alarm" and "Report" in Alarm_report_path:
                msg = 'File is not a Daily Alarm Report'
                sg.popup(msg, title="Error with the selections")
                # print('File is not a Daily Alarm Report')
            elif not "Irradiance" in irradiance_file_path:
                msg = 'File is not a Irradiance file'
                sg.popup(msg, title="Error with the selections")
                # print('File is not a Daily Alarm Report')
    # window.close()

    return


def dmrprocess2(incidents_file="No File", tracker_incidents_file="No File",
                site_list=["No site list"], geography="PT", date="27-03-1996"):
    sg.theme('DarkAmber')  # Add a touch of color
    if incidents_file == "No File" or tracker_incidents_file == "No File":
        sg.popup('No files or site list available, please select them', no_titlebar=True)
        incidents_file, tracker_incidents_file, site_list, geography, date = inputs.choose_incidents_files()
        if incidents_file == "No File" or tracker_incidents_file == "No File":
            return None
    else:
        print("Incidents file: " + incidents_file + "\nTracker Incidents file: " + tracker_incidents_file)

    dir = os.path.dirname(incidents_file)
    reportfiletemplate = dir + '/Info&Templates/Reporting_' + geography + '_Sites_' + 'Template.xlsx'
    general_info_path = dir + '/Info&Templates/General Info ' + geography + '.xlsx'

    # Reset Report Template to create new report
    reportfile_path = data_treatment.reset_final_report(reportfiletemplate, date, geography)

    # Read Active and Closed Events
    df_list_active, df_list_closed = data_acquisition.read_approved_incidents(incidents_file, site_list, roundto=1)
    df_tracker_active, df_tracker_closed = data_acquisition.read_approved_tracker_inc(tracker_incidents_file, roundto=1)

    # Read sunrise and sunset hours
    df_info_sunlight = pd.read_excel(incidents_file, sheet_name='Info', engine="openpyxl")
    df_info_sunlight['Time of operation start'] = df_info_sunlight['Time of operation start'].dt.round(freq='s')
    df_info_sunlight['Time of operation end'] = df_info_sunlight['Time of operation end'].dt.round(freq='s')

    # Describe Incidents
    df_list_active = data_treatment.describe_incidents(df_list_active, df_info_sunlight, active_events=True,
                                                       tracker=False)
    df_list_closed = data_treatment.describe_incidents(df_list_closed, df_info_sunlight, active_events=False,
                                                       tracker=False)
    df_tracker_active = data_treatment.describe_incidents(df_tracker_active, df_info_sunlight, active_events=True,
                                                          tracker=True)
    df_tracker_closed = data_treatment.describe_incidents(df_tracker_closed, df_info_sunlight, active_events=False,
                                                          tracker=True)
    print(df_tracker_closed.columns)

    # Add Events to Report File
    add_events_to_final_report(reportfile_path, df_list_active, df_list_closed, df_tracker_active, df_tracker_closed)

    # -------------------------------Analysis on Components Failures---------------------------------
    # Read and update the timestamps on the analysis dataframes
    df_incidents_analysis, df_tracker_analysis = data_treatment.read_analysis_df_and_correct_date(reportfiletemplate,
                                                                                                  date, roundto=1)
    # Analysis of components failures
    df_incidents_analysis_final = data_analysis.analysis_component_incidents(
        df_incidents_analysis, site_list, df_list_closed, df_list_active, df_info_sunlight)

    # Analysis of tracker failures
    df_tracker_analysis_final = data_analysis.analysis_tracker_incidents(
        df_tracker_analysis, df_tracker_closed, df_tracker_active, df_info_sunlight)

    # Add Analysis to excel file
    add_analysis_to_reportfile(reportfile_path, df_incidents_analysis_final, df_tracker_analysis_final,
                               df_info_sunlight)

    return reportfile_path


def dmrprocess2_new(incidents_file="No File", tracker_incidents_file="No File", site_list=["No site list"],
                    geography="PT", date="1996-03-27"):
    period = "day"
    sg.theme('DarkAmber')  # Add a touch of color
    if incidents_file == "No File" or tracker_incidents_file == "No File":
        sg.popup('No files or site list available, please select them', no_titlebar=True)
        incidents_file, tracker_incidents_file, site_list, geography, date = inputs.choose_incidents_files()
        if incidents_file == "No File" or tracker_incidents_file == "No File":
            return None
    else:
        print("Incidents file: " + incidents_file + "\nTracker Incidents file: " + tracker_incidents_file)

    main_dir = os.path.dirname(incidents_file)
    username = os.getlogin()

    dest_file = main_dir + '//Reporting_' + geography + '_Sites_' + date.replace("-", "") + "_" + username + '.xlsx'
    general_info_path = main_dir + '/Info&Templates/General Info ' + geography + '.xlsx'
    irradiance_file_path = main_dir + '/Irradiance ' + geography + '/Irradiance_' + geography + '_Curated&Average-' + \
                           date.replace("-", "") + '.xlsx'
    export_file_path = main_dir + '/Exported Energy ' + geography + '/Energy_Exported_' + geography + '_' + \
                       date.replace("-", "") + '.xlsx'

    # Read irradiance and export files
    irradiance_df, export_df = data_acquisition.read_irradiance_export(irradiance_file_path, export_file_path)

    # Read general info files
    component_data, tracker_data, fmeca_data, site_capacities, fleet_capacity, budget_irr, budget_pr, budget_export, \
    all_site_info = data_acquisition.get_general_info_dataframes(general_info_path)

    # Read Active and Closed Events
    df_list_active, df_list_closed = data_acquisition.read_approved_incidents(incidents_file, site_list, roundto=1)
    df_tracker_active, df_tracker_closed = data_acquisition.read_approved_tracker_inc(tracker_incidents_file, roundto=1)

    # Read sunrise and sunset hours
    df_info_sunlight = pd.read_excel(incidents_file, sheet_name='Info', engine="openpyxl")
    df_info_sunlight['Time of operation start'] = df_info_sunlight['Time of operation start'].dt.round(freq='s')
    df_info_sunlight['Time of operation end'] = df_info_sunlight['Time of operation end'].dt.round(freq='s')

    # <editor-fold desc="Describe Incidents">
    df_list_active = data_treatment.describe_incidents(df_list_active, df_info_sunlight, active_events=True,
                                                       tracker=False)
    df_list_closed = data_treatment.describe_incidents(df_list_closed, df_info_sunlight, active_events=False,
                                                       tracker=False)
    df_tracker_active = data_treatment.describe_incidents(df_tracker_active, df_info_sunlight, active_events=True,
                                                          tracker=True)
    df_tracker_closed = data_treatment.describe_incidents(df_tracker_closed, df_info_sunlight, active_events=False,
                                                          tracker=True)

    # </editor-fold>

    # <editor-fold desc="Complete incidents list">
    df_active = data_treatment.match_df_to_event_tracker(pd.concat(df_list_active.values(), ignore_index=True),
                                                         component_data, fmeca_data, active=True)

    df_closed = data_treatment.match_df_to_event_tracker(pd.concat(df_list_closed.values(), ignore_index=True),
                                                         component_data, fmeca_data)

    df_tracker_active = data_treatment.match_df_to_event_tracker(df_tracker_active, tracker_data, fmeca_data,
                                                                 active=True, tracker=True)

    df_tracker_closed = data_treatment.match_df_to_event_tracker(df_tracker_closed, tracker_data, fmeca_data,
                                                                 tracker=True)

    final_df_to_add = {'Active incidents': df_active,
                       "Closed incidents": df_closed,
                       "Active tracker incidents": df_tracker_active,
                       "Closed tracker incidents": df_tracker_closed}

    # </editor-fold>

    # <editor-fold desc="Create concatenated incidents list of incidents of the day">
    df_active["Event End Time"] = ""
    df_tracker_active["Event End Time"] = ""

    df_incidents = pd.concat([df_active, df_closed])

    # print(df_incidents[["Related Component", "Event Start Time", "Event End Time"]])

    df_incidents["Event Start Time"] = [datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S') for timestamp
                                        in df_incidents["Event Start Time"]]

    df_incidents["Event End Time"] = [datetime.strptime(str(date + " 23:00:00"), '%Y-%m-%d %H:%M:%S')
                                      if timestamp == "" else datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S') for
                                      timestamp in df_incidents["Event End Time"]]

    # </editor-fold>

    # Availability Calculation

    availability_period_df, raw_availability_period_df, activehours_period_df, incidents_corrected_period, \
    all_corrected_incidents, date_range = calculations.availability_in_period(df_incidents, period, component_data,
                                                                              irradiance_df, export_df, budget_pr,
                                                                              irradiance_threshold=20, timestamp=15,
                                                                              date=date)

    # print(availability_period_df)
    final_df_to_add["Incidents Daily Overview"] = data_treatment.match_df_to_event_tracker(all_corrected_incidents,
                                                                                           component_data, fmeca_data,
                                                                                           simple_match=True)

    pr_data_period_df = calculations.pr_in_period(all_corrected_incidents, availability_period_df,
                                                  raw_availability_period_df, period, component_data, irradiance_df,
                                                  export_df, budget_pr, budget_export, budget_irr,
                                                  irradiance_threshold=20, timestamp=15, date=date)

    pr_data_period_df = calculations.day_end_availability(pr_data_period_df, final_df_to_add, component_data,
                                                          tracker_data, all_site_info)

    print(pr_data_period_df.iloc[:4, :2])

    create_dmr_file(final_df_to_add, dest_file, pr_data_period_df, site_capacities)

    return dest_file


def create_dmr_file(final_df_to_add, dest_file, performance_fleet_period, site_capacities):
    writer = pd.ExcelWriter(dest_file, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}})
    workbook = writer.book

    # <editor-fold desc="Formats">
    # Format column header
    format_darkblue_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#002060', 'font_color': '#FFFFFF'})
    format_darkblue_white.set_bold()
    format_darkblue_white.set_text_wrap()

    format_lightblue_black = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'font_color': '#000000'})
    format_lightblue_black.set_bold()
    format_lightblue_black.set_text_wrap()
    format_lightblue_black.set_border()

    format_header = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#D9D9D9', 'font_color': '#000000'})
    format_header.set_bold()
    format_header.set_text_wrap()

    format_all_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFFFFF', 'font_color': '#FFFFFF'})
    format_all_black = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#000000', 'font_color': '#000000'})
    format_black_on_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#000000', 'font_color': '#FFFFFF'})

    # Format of specific column data
    format_day_data = workbook.add_format({'num_format': 'dd/mm/yyyy', 'valign': 'vcenter'})
    format_day_data.set_align('right')
    format_day_data.set_border()

    format_hour_data = workbook.add_format({'num_format': 'hh:mm:ss', 'valign': 'vcenter'})
    format_hour_data.set_align('right')
    format_hour_data.set_border()

    format_day_hour = workbook.add_format({'num_format': 'dd/mm/yyyy hh:mm:ss', 'valign': 'vcenter'})
    format_day_hour.set_align('right')
    format_day_hour.set_border()

    # Format numbers
    format_number = workbook.add_format({'num_format': '#,##0.00', 'align': 'center', 'valign': 'vcenter'})
    format_number.set_border()

    format_nodecimal = workbook.add_format({'num_format': '0', 'align': 'center', 'valign': 'vcenter'})
    format_nodecimal.set_border()

    format_percentage = workbook.add_format({'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter'})
    format_percentage.set_border()

    format_percentage_good = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#C6EFCE',
         'font_color': '#006100'})
    format_percentage_good.set_border()
    format_percentage_mid = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFEB9C',
         'font_color': '#9C5700'})
    format_percentage_mid.set_border()
    format_percentage_bad = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFC7CE',
         'font_color': '#9C0006'})
    format_percentage_bad.set_border()

    # Format strings
    format_string = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
    format_string.set_border()

    format_string_wrapped = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
    format_string_wrapped.set_text_wrap()
    format_string_wrapped.set_border()

    format_string_unlocked = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'locked': False})
    unlocked = workbook.add_format({'locked': False})
    format_string_unlocked.set_border()

    format_string_bold = workbook.add_format({'align': 'right', 'valign': 'vcenter'})
    format_string_bold.set_bold()
    format_string_bold.set_border()

    format_string_bold_wrapped = workbook.add_format({'align': 'right', 'valign': 'vcenter'})
    format_string_bold_wrapped.set_bold()
    format_string_bold_wrapped.set_border()
    format_string_bold_wrapped.set_text_wrap()

    format_first_column = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#F2F2F2', 'font_color': '#000000'})
    format_first_column.set_bold()
    format_first_column.set_border()
    format_first_column.set_text_wrap()
    # </editor-fold>

    # <editor-fold desc="Performance Overview Sheet">
    active_events = final_df_to_add['Active incidents']

    overview_events = active_events.loc[active_events['Component Status'] == "Not Producing"][
        ['Site Name', 'ID', 'Related Component', 'Event Start Time', 'Energy Lost (MWh)', 'Capacity Related Component']]

    overview_events['% of site affected'] = ["{:.2%}".format(row['Capacity Related Component'] /
                                                             float(site_capacities.loc[row['Site Name']]))
                                             for index, row in overview_events.iterrows()]
    # overview_events

    sheet = "Performance Overview"
    try:
        ws_sheet = workbook.add_worksheet(sheet)
    except (xlsxwriter.exceptions.DuplicateWorksheetName, NameError):
        sheet = sheet + "_new"
        ws_sheet = workbook.add_worksheet(sheet)

    df_performance = performance_fleet_period.T

    # print(df_performance)

    sites = list(df_performance.columns)

    start_row_header = 1
    start_row_data = 2
    start_column = 0

    for site in sites:
        level = 0
        start_row_header_str = str(start_row_header)
        start_row_data_str = str(start_row_data)

        performance_site = df_performance.loc[:, [site]].reset_index()
        print(performance_site)
        n_columns_total = performance_site.shape[1]

        width = get_col_widths(performance_site)

        # print("\n", df_total)

        for i in range(start_column, start_column + n_columns_total):

            header = performance_site.columns[i - start_column]

            column_letter = openpyxl.utils.cell.get_column_letter(i + 1)
            header_cell = column_letter + start_row_header_str
            data_cell = column_letter + start_row_data_str
            all_column = column_letter + ':' + column_letter

            # print('Header: ', header, "\n", 'Header cell:', header_cell, "\n", "Data Cell: ", data_cell ,"\n")

            data = list(performance_site[header].fillna(""))

            if header == "index":
                to_collapse_column1 = column_letter
                data = [x for x in data if not pd.isnull(x)]

                if column_letter == "A":
                    ws_sheet.write(header_cell, "", format_darkblue_white)
                    ws_sheet.write_column(data_cell, data, format_lightblue_black)
                    ws_sheet.set_column(all_column, 23)
                else:
                    pass

            elif header in sites:
                kpis = performance_site['index']
                ws_sheet.write(header_cell, header, format_darkblue_white)
                data = [x for x in data if not x == ""]

                for i in range(len(data)):
                    cell = column_letter + str(start_row_data + i)
                    value = data[i]
                    kpi = kpis[i]

                    if "%" in value:
                        value = float(value[:-1]) / 100
                        ws_sheet.write_number(cell, value, format_percentage)
                        if not "PR (%)" in kpi:
                            if "Availability" in kpi:
                                ws_sheet.conditional_format(cell, {'type': 'icon_set', 'icon_style': '3_traffic_lights',
                                                                   'icons': [{'criteria': '>=', 'type': 'number',
                                                                              'value': 0.97},
                                                                             {'criteria': '<', 'type': 'number',
                                                                              'value': 0.10},
                                                                             {'criteria': '<=', 'type': 'number',
                                                                              'value': 0.10}]})

                            else:
                                ws_sheet.conditional_format(cell, {'type': 'icon_set', 'icon_style': '3_traffic_lights',
                                                                   'icons': [
                                                                       {'criteria': '<=', 'type': 'number', 'value': 0},
                                                                       {'criteria': '>', 'type': 'number', 'value': 0},
                                                                       {'criteria': '>=', 'type': 'number',
                                                                        'value': 0.05}]})

                    elif "Portfolio" or "Fault Status" in kpi:
                        ws_sheet.write(cell, value, format_string)

                    else:
                        value = float(value.replace(",", ""))
                        ws_sheet.write_number(cell, value, format_number)

                ws_sheet.set_column(all_column, 16, None)

            else:
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_string)
                ws_sheet.set_column(all_column, 18, None, {'level': 1, 'hidden': True})

        # level = level + 1
        start_column = start_column + n_columns_total - 1
    # </editor-fold>

    # <editor-fold desc="Incidents' sheets">

    for sheet in final_df_to_add.keys():
        df = final_df_to_add[sheet]
        width = get_col_widths(df)
        n_rows = df.shape[0]
        n_columns = df.shape[1]
        try:
            ws_sheet = workbook.add_worksheet(sheet)
        except (xlsxwriter.exceptions.DuplicateWorksheetName, NameError):
            sheet = sheet + "_new"
            ws_sheet = workbook.add_worksheet(sheet)

        for i in range(len(df.columns)):
            header = df.columns[i]
            column_letter = openpyxl.utils.cell.get_column_letter(i + 1)
            header_cell = column_letter + '1'
            data_cell = column_letter + '2'
            all_column = column_letter + ':' + column_letter
            data = df[header].fillna("")
            # print(data)

            if header == 'ID':
                ws_sheet.write(header_cell, header, format_header)
                ws_sheet.write_column(data_cell, data, format_first_column)
                ws_sheet.set_column(all_column, 18)

            elif "Time" in header:
                ws_sheet.write(header_cell, header, format_header)
                ws_sheet.write_column(data_cell, data, format_day_hour)
                ws_sheet.set_column(all_column, 19)

            elif "Capacity" in header or "(" in header:
                ws_sheet.write(header_cell, header, format_header)
                ws_sheet.write_column(data_cell, data, format_number)
                ws_sheet.set_column(all_column, 12)


            elif "Fa" in header or "ategory" in header or "Excludable" in header:
                if header == "Resolution Category":
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_string_unlocked)
                    ws_sheet.set_column(all_column, width[i + 1], unlocked)
                    ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                             {'validate': 'list', 'source': ['Repair',
                                                                             'Reset',
                                                                             'Part Replacement',
                                                                             'Unit Replacement']})

                elif "Excludable" in header:
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_string_unlocked)
                    ws_sheet.set_column(all_column, width[i + 1], unlocked)
                    # ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                    #  {'validate': 'list',
                    #   'source': ['OMC', 'Force Majeure', 'Curtailment', "N/A"]})

                else:
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_string)
                    ws_sheet.set_column(all_column, width[i + 1])


            elif header == 'Remediation' or header == 'Comments':
                ws_sheet.write(header_cell, header, format_header)
                ws_sheet.write_column(data_cell, data, format_string_wrapped)
                ws_sheet.set_column(all_column, 60)

            else:
                ws_sheet.write(header_cell, header, format_header)
                ws_sheet.write_column(data_cell, data, format_string)
                ws_sheet.set_column(all_column, width[i + 1])

        ws_sheet.set_default_row(30)
    # </editor-fold>

    ws_active = workbook.get_worksheet_by_name("Performance Overview")
    ws_active.activate()

    writer.save()
    print('Done')

    return


def create_curtailment_file(dest_file, site_selection, curtailment_events_by_site, monthly_curtailment_by_site,
                            component_data, fmeca_data):
    writer = pd.ExcelWriter(dest_file, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}})
    workbook = writer.book

    # <editor-fold desc="Formats">
    # Format column header
    format_darkblue_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#002060', 'font_color': '#FFFFFF'})
    format_darkblue_white.set_bold()
    format_darkblue_white.set_text_wrap()

    format_lightblue_black = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'font_color': '#000000'})
    format_lightblue_black.set_bold()
    format_lightblue_black.set_text_wrap()
    format_lightblue_black.set_border()

    format_header = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#D9D9D9', 'font_color': '#000000'})
    format_header.set_bold()
    format_header.set_text_wrap()

    format_all_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFFFFF', 'font_color': '#FFFFFF'})
    format_all_black = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#000000', 'font_color': '#000000'})
    format_black_on_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#000000', 'font_color': '#FFFFFF'})

    # Format of specific column data
    format_day_data = workbook.add_format({'num_format': 'dd/mm/yyyy', 'valign': 'vcenter'})
    format_day_data.set_align('right')
    format_day_data.set_border()

    format_hour_data = workbook.add_format({'num_format': 'hh:mm:ss', 'valign': 'vcenter'})
    format_hour_data.set_align('right')
    format_hour_data.set_border()

    format_day_hour = workbook.add_format({'num_format': 'dd/mm/yyyy hh:mm:ss', 'valign': 'vcenter'})
    format_day_hour.set_align('right')
    format_day_hour.set_border()

    # Format numbers
    format_number = workbook.add_format({'num_format': '#,##0.00', 'align': 'center', 'valign': 'vcenter'})
    format_number.set_border()

    format_nodecimal = workbook.add_format({'num_format': '0', 'align': 'center', 'valign': 'vcenter'})
    format_nodecimal.set_border()

    format_percentage = workbook.add_format({'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter'})
    format_percentage.set_border()

    format_percentage_good = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#C6EFCE', 'font_color': '#006100'})
    format_percentage_good.set_border()
    format_percentage_mid = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFEB9C', 'font_color': '#9C5700'})
    format_percentage_mid.set_border()
    format_percentage_bad = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
    format_percentage_bad.set_border()

    # Format strings
    format_string = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
    format_string.set_border()

    format_string_wrapped = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
    format_string_wrapped.set_text_wrap()
    format_string_wrapped.set_border()

    format_string_unlocked = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'locked': False})
    unlocked = workbook.add_format({'locked': False})
    format_string_unlocked.set_border()

    format_string_bold = workbook.add_format({'align': 'right', 'valign': 'vcenter'})
    format_string_bold.set_bold()
    format_string_bold.set_border()

    format_string_bold_wrapped = workbook.add_format({'align': 'right', 'valign': 'vcenter'})
    format_string_bold_wrapped.set_bold()
    format_string_bold_wrapped.set_border()
    format_string_bold_wrapped.set_text_wrap()

    format_first_column = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#F2F2F2', 'font_color': '#000000'})
    format_first_column.set_bold()
    format_first_column.set_border()
    # </editor-fold>

    for site in site_selection:

        sheet_raw_df = site.replace("LSBP - ", "") + " raw DF"
        sheet_incidents = site.replace("LSBP - ", "") + " events"
        sheet_month = site.replace("LSBP - ", "") + " MO"

        # Add raw df
        df_site = curtailment_events_by_site[site]
        # df_site.to_excel(writer, sheet_name=sheet_raw_df)

        # Add incidents df
        ws_sheet = workbook.add_worksheet(sheet_incidents)
        et_match_df = data_treatment.match_df_to_event_tracker(df_site, component_data, fmeca_data)

        width = get_col_widths(et_match_df)

        for i in range(len(et_match_df.columns)):
            header = et_match_df.columns[i]
            column_letter = openpyxl.utils.cell.get_column_letter(i + 1)
            header_cell = column_letter + '1'
            data_cell = column_letter + '2'
            all_column = column_letter + ':' + column_letter
            data = et_match_df[header].fillna("")

            if header == 'ID':
                ws_sheet.write(header_cell, header, format_header)
                ws_sheet.write_column(data_cell, data, format_first_column)
                ws_sheet.set_column(all_column, width[i + 1])

            elif "Time" in header:
                ws_sheet.write(header_cell, header, format_header)
                ws_sheet.write_column(data_cell, data, format_day_hour)
                ws_sheet.set_column(all_column, width[i + 1])

            elif "Capacity" in header or "(" in header:
                ws_sheet.write(header_cell, header, format_header)
                ws_sheet.write_column(data_cell, data, format_number)
                ws_sheet.set_column(all_column, width[i + 1])



            elif header == "Incident Status":
                ws_sheet.write(header_cell, header, format_header)
                ws_sheet.write_column(data_cell, data, format_string_unlocked)
                ws_sheet.set_column(all_column, width[i + 1], unlocked)

            elif header == "Categorization Status":
                ws_sheet.write(header_cell, header, format_header)
                ws_sheet.write_column(data_cell, data, format_string_unlocked)
                ws_sheet.set_column(all_column, width[i + 1], unlocked)


            elif header == 'Remediation' or header == 'Comments':
                ws_sheet.write(header_cell, header, format_header)
                ws_sheet.write_column(data_cell, data, format_string_wrapped)
                ws_sheet.set_column(all_column, 80)


            else:
                ws_sheet.write(header_cell, header, format_header)
                ws_sheet.write_column(data_cell, data, format_string)
                ws_sheet.set_column(all_column, width[i + 1])

        # Add monthly sheet
        df_site_month = monthly_curtailment_by_site[site]
        df_site_month.to_excel(writer, sheet_name=sheet_month)

    writer.close()

    writer.handles = None

    print('Done')

    # IPython.Application.instance().kernel.do_shutdown(True)

    return


def create_clipping_file(site, summaries_site, dest_file, graphs_site):
    dest_file = dest_file.replace(".xlsx", "_" + site.replace("LSBP - ", "") + ".xlsx")

    writer = pd.ExcelWriter(dest_file, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}})
    workbook = writer.book

    # <editor-fold desc="Formats">
    # Format column header
    format_darkblue_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#002060', 'font_color': '#FFFFFF'})
    format_darkblue_white.set_bold()
    format_darkblue_white.set_text_wrap()

    format_lightblue_black = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'font_color': '#000000'})
    format_lightblue_black.set_bold()
    format_lightblue_black.set_text_wrap()
    format_lightblue_black.set_border()

    format_header = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#D9D9D9', 'font_color': '#000000'})
    format_header.set_bold()
    format_header.set_text_wrap()

    format_all_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFFFFF', 'font_color': '#FFFFFF'})
    format_all_black = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#000000', 'font_color': '#000000'})
    format_black_on_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#000000', 'font_color': '#FFFFFF'})

    # Format of specific column data
    format_day_data = workbook.add_format({'num_format': 'dd/mm/yyyy', 'valign': 'vcenter'})
    format_day_data.set_align('right')
    format_day_data.set_border()

    format_hour_data = workbook.add_format({'num_format': 'hh:mm:ss', 'valign': 'vcenter'})
    format_hour_data.set_align('right')
    format_hour_data.set_border()

    format_day_hour = workbook.add_format({'num_format': 'dd/mm/yyyy hh:mm:ss', 'valign': 'vcenter'})
    format_day_hour.set_align('right')
    format_day_hour.set_border()

    # Format numbers
    format_number = workbook.add_format({'num_format': '#,##0.00', 'align': 'center', 'valign': 'vcenter'})
    format_number.set_border()

    format_nodecimal = workbook.add_format({'num_format': '0', 'align': 'center', 'valign': 'vcenter'})
    format_nodecimal.set_border()

    format_percentage = workbook.add_format({'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter'})
    format_percentage.set_border()

    format_percentage_good = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#C6EFCE', 'font_color': '#006100'})
    format_percentage_good.set_border()
    format_percentage_mid = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFEB9C', 'font_color': '#9C5700'})
    format_percentage_mid.set_border()
    format_percentage_bad = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
    format_percentage_bad.set_border()

    # Format strings
    format_string = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
    format_string.set_border()

    format_string_wrapped = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
    format_string_wrapped.set_text_wrap()
    format_string_wrapped.set_border()

    format_string_unlocked = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'locked': False})
    unlocked = workbook.add_format({'locked': False})
    format_string_unlocked.set_border()

    format_string_bold = workbook.add_format({'align': 'right', 'valign': 'vcenter'})
    format_string_bold.set_bold()
    format_string_bold.set_border()

    format_string_bold_wrapped = workbook.add_format({'align': 'right', 'valign': 'vcenter'})
    format_string_bold_wrapped.set_bold()
    format_string_bold_wrapped.set_border()
    format_string_bold_wrapped.set_text_wrap()

    format_first_column = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#F2F2F2', 'font_color': '#000000'})
    format_first_column.set_bold()
    format_first_column.set_border()
    # </editor-fold>

    for key in summaries_site.keys():
        sheet = site + "_" + key
        sheet = sheet.replace("LSBP - ", "")

        try:
            ws_sheet = workbook.add_worksheet(sheet)
        except (xlsxwriter.exceptions.DuplicateWorksheetName, NameError):
            sheet = sheet + "_new"
            ws_sheet = workbook.add_worksheet(sheet)

        df = summaries_site[key]
        width = get_col_widths(df)
        n_rows = df.shape[0]
        n_columns = df.shape[1]
        index = df.index
        index_name = df.index.name

        ws_sheet.write("A1", index_name, format_header)
        ws_sheet.write_column("A2", index, format_day_data)
        ws_sheet.set_column("A:A", 15)

        for i in range(len(df.columns)):
            header = df.columns[i]
            column_letter = openpyxl.utils.cell.get_column_letter(i + 2)
            header_cell = column_letter + '1'
            data_cell = column_letter + '2'
            all_column = column_letter + ':' + column_letter
            data = df[header].fillna("")

            if "%" in header:
                ws_sheet.write(header_cell, header, format_header)
                ws_sheet.write_column(data_cell, data, format_number)
                ws_sheet.set_column(all_column, width[i + 1])

            else:
                ws_sheet.write(header_cell, header, format_header)
                ws_sheet.write_column(data_cell, data, format_number)
                ws_sheet.set_column(all_column, width[i + 1])

        ws_sheet.set_default_row(30)

        # Insert Images
        image_column = openpyxl.utils.cell.get_column_letter(i + 3)
        image_column2 = openpyxl.utils.cell.get_column_letter(i + 23)
        image_row_n = 1

        graphs_energy = graphs_site['Energy']
        graphs_loss = graphs_site['% of loss']

        # for site_key in graphs_site.keys():

        graph_energy_granularity = graphs_energy[key]
        graph_loss_granularity = graphs_loss[key]

        image_cell = image_column + str(image_row_n)
        image_cell2 = image_column2 + str(image_row_n)

        ws_sheet.insert_image(image_cell, graph_energy_granularity)
        ws_sheet.insert_image(image_cell2, graph_loss_granularity)

        image_row_n = image_row_n + 10

        ws_sheet.set_column((image_column + ':ZZ'), 10, format_all_white)

    writer.close()

    writer.handles = None

    print('Done')

    return


# <editor-fold desc="ET Functions">


# <editor-fold desc="Xlsxwriter related custom functions">

def get_rowindex_and_columnletter(cell):
    cell_letter_code = re.search(r'([\w]+)([\d]+)', cell)
    rowindex = cell_letter_code.group(0)
    column_letter = cell_letter_code.group(1)
    return rowindex, column_letter


def get_col_widths(dataframe):
    # First we find the maximum length of the index column
    idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]


# </editor-fold>

def create_event_tracker_file_all(final_df_to_add, dest_file, performance_fleet_per_period, site_capacities,
                                  dict_fmeca_shapes):
    writer = pd.ExcelWriter(dest_file, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}})
    workbook = writer.book

    # <editor-fold desc="Formats">
    # Format column header
    format_darkblue_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#002060', 'font_color': '#FFFFFF'})
    format_darkblue_white.set_bold()
    format_darkblue_white.set_text_wrap()

    format_lightblue_black = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'font_color': '#000000'})
    format_lightblue_black.set_bold()
    format_lightblue_black.set_text_wrap()
    format_lightblue_black.set_border()

    format_header = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#D9D9D9', 'font_color': '#000000'})
    format_header.set_bold()
    format_header.set_text_wrap()

    format_all_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFFFFF', 'font_color': '#FFFFFF'})
    format_all_black = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#000000', 'font_color': '#000000'})
    format_black_on_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#000000', 'font_color': '#FFFFFF'})

    # Format of specific column data
    format_day_data = workbook.add_format({'num_format': 'dd/mm/yyyy', 'valign': 'vcenter'})
    format_day_data.set_align('right')
    format_day_data.set_border()

    format_hour_data = workbook.add_format({'num_format': 'hh:mm:ss', 'valign': 'vcenter'})
    format_hour_data.set_align('right')
    format_hour_data.set_border()

    format_day_hour = workbook.add_format({'num_format': 'dd/mm/yyyy hh:mm:ss', 'valign': 'vcenter'})
    format_day_hour.set_align('right')
    format_day_hour.set_border()

    # Format numbers
    format_number = workbook.add_format({'num_format': '#,##0.00', 'align': 'center', 'valign': 'vcenter'})
    format_number.set_border()

    format_nodecimal = workbook.add_format({'num_format': '0', 'align': 'center', 'valign': 'vcenter'})
    format_nodecimal.set_border()

    format_percentage = workbook.add_format({'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter'})
    format_percentage.set_border()

    format_percentage_good = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#C6EFCE',
         'font_color': '#006100'})
    format_percentage_good.set_border()
    format_percentage_mid = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFEB9C',
         'font_color': '#9C5700'})
    format_percentage_mid.set_border()
    format_percentage_bad = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFC7CE',
         'font_color': '#9C0006'})
    format_percentage_bad.set_border()

    # Format strings
    format_string = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
    format_string.set_border()

    format_string_wrapped = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
    format_string_wrapped.set_text_wrap()
    format_string_wrapped.set_border()

    format_string_unlocked = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'locked': False})
    unlocked = workbook.add_format({'locked': False})
    format_string_unlocked.set_border()

    format_string_bold = workbook.add_format({'align': 'right', 'valign': 'vcenter'})
    format_string_bold.set_bold()
    format_string_bold.set_border()

    format_string_bold_wrapped = workbook.add_format({'align': 'right', 'valign': 'vcenter'})
    format_string_bold_wrapped.set_bold()
    format_string_bold_wrapped.set_border()
    format_string_bold_wrapped.set_text_wrap()

    format_first_column = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#F2F2F2', 'font_color': '#000000'})
    format_first_column.set_bold()
    format_first_column.set_border()
    format_first_column.set_text_wrap()
    # </editor-fold>

    # <editor-fold desc="YTD Performance Overview Sheet">
    sheet = "YTD Performance Overview"
    try:
        ws_sheet = workbook.add_worksheet(sheet)
    except (xlsxwriter.exceptions.DuplicateWorksheetName, NameError):
        sheet = sheet + "_new"
        ws_sheet = workbook.add_worksheet(sheet)

    df_performance = performance_fleet_per_period['ytd'].T

    sites = list(df_performance.columns)

    start_row_header = 1
    start_row_data = 2
    start_column = 0

    for site in sites:
        level = 0
        start_row_header_str = str(start_row_header)
        start_row_data_str = str(start_row_data)

        performance_site = df_performance.loc[:, [site]].reset_index()
        n_rows_performance = performance_site.shape[0] + 1
        n_columns_performance = performance_site.shape[1]

        df_total = performance_site

        max_rows = n_rows_performance
        n_columns_total = df_total.shape[1]

        width = get_col_widths(df_total)

        print("\n", df_total)

        for i in range(start_column, start_column + n_columns_total):

            header = df_total.columns[i - start_column]

            column_letter = openpyxl.utils.cell.get_column_letter(i + 1)
            header_cell = column_letter + start_row_header_str
            data_cell = column_letter + start_row_data_str
            all_column = column_letter + ':' + column_letter

            # print('Header: ', header, "\n", 'Header cell:', header_cell, "\n", "Data Cell: ", data_cell ,"\n")

            data = list(df_total[header].fillna(""))

            if header == "index":
                to_collapse_column1 = column_letter
                data = [x for x in data if not pd.isnull(x)]
                ws_sheet.write(header_cell, "", format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_lightblue_black)
                if column_letter == "A":
                    ws_sheet.set_column(all_column, 23)
                else:
                    ws_sheet.set_column(all_column, 0)  # ,None,{'level': 1, 'hidden': True})

            elif "LSBP" in header or "Wellington" in header:
                kpis = df_total['index']
                ws_sheet.write(header_cell, header, format_darkblue_white)
                data = [x for x in data if not x == ""]

                for i in range(len(data)):
                    cell = column_letter + str(start_row_data + i)
                    value = data[i]
                    kpi = kpis[i]

                    if "%" in value:
                        value = float(value[:-1]) / 100
                        ws_sheet.write_number(cell, value, format_percentage)
                        if not "PR (%)" in kpi:
                            if "Availability" in kpi:
                                ws_sheet.conditional_format(cell, {'type': 'icon_set', 'icon_style': '3_traffic_lights',
                                                                   'icons': [{'criteria': '>=', 'type': 'number',
                                                                              'value': 0.97},
                                                                             {'criteria': '<', 'type': 'number',
                                                                              'value': 0.10},
                                                                             {'criteria': '<=', 'type': 'number',
                                                                              'value': 0.10}]})

                            else:
                                ws_sheet.conditional_format(cell, {'type': 'icon_set', 'icon_style': '3_traffic_lights',
                                                                   'icons': [
                                                                       {'criteria': '<=', 'type': 'number', 'value': 0},
                                                                       {'criteria': '>', 'type': 'number', 'value': 0},
                                                                       {'criteria': '>=', 'type': 'number',
                                                                        'value': 0.05}]})


                    else:
                        value = float(value.replace(",", ""))
                        ws_sheet.write_number(cell, value, format_number)

                ws_sheet.set_column(all_column, 16, None)


            else:
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_string)
                ws_sheet.set_column(all_column, 18, None)  # ,{'level': 1, 'hidden': True})

        level = level + 1
        start_column = start_column + n_columns_total
    # </editor-fold>

    # <editor-fold desc="MTD Performance Overview Sheet">
    active_events = final_df_to_add['Active Events']
    overview_events = active_events.loc[active_events['Component Status'] == "Not Producing"][
        ['Site Name', 'ID', 'Related Component', 'Event Start Time', 'Energy Lost (MWh)', 'Capacity Related Component']]
    overview_events['% of site affected'] = [
        "{:.2%}".format(row['Capacity Related Component'] / float(site_capacities.loc[row['Site Name']])) for index, row
        in overview_events.iterrows()]
    overview_events['Actions'] = active_events.loc[active_events['Component Status'] == "Not Producing"]['Remediation']
    overview_events['Space'] = ""
    # overview_events

    sheet = "MTD Performance Overview"
    try:
        ws_sheet = workbook.add_worksheet(sheet)
    except (xlsxwriter.exceptions.DuplicateWorksheetName, NameError):
        sheet = sheet + "_new"
        ws_sheet = workbook.add_worksheet(sheet)

    try:
        df_performance = performance_fleet_per_period['mtd'].T
    except KeyError:
        df_performance = performance_fleet_per_period['monthly'].T

    sites = list(df_performance.columns)

    start_row_header = 1
    start_row_data = 2
    start_column = 0

    for site in sites:
        level = 0
        start_row_header_str = str(start_row_header)
        start_row_data_str = str(start_row_data)

        performance_site = df_performance.loc[:, [site]].reset_index()
        n_rows_performance = performance_site.shape[0] + 1
        n_columns_performance = performance_site.shape[1]

        incidents_site = overview_events.loc[overview_events['Site Name'] == site].reset_index(drop=True)
        # incidents_site.insert(1, "#", list(range(1,incidents_site.shape[0] + 1)))
        n_rows_incidents = incidents_site.shape[0] + 1
        n_columns_incidents = incidents_site.shape[1]

        df_total = pd.concat([performance_site, incidents_site], axis=1)

        max_rows = max(n_rows_performance, n_rows_incidents)
        n_columns_total = df_total.shape[1]

        width = get_col_widths(df_total)

        # print("\n", df_total)

        for i in range(start_column, start_column + n_columns_total):

            header = df_total.columns[i - start_column]

            column_letter = openpyxl.utils.cell.get_column_letter(i + 1)
            header_cell = column_letter + start_row_header_str
            data_cell = column_letter + start_row_data_str
            all_column = column_letter + ':' + column_letter

            # print('Header: ', header, "\n", 'Header cell:', header_cell, "\n", "Data Cell: ", data_cell ,"\n")

            data = list(df_total[header].fillna(""))

            if header == "index":
                to_collapse_column1 = column_letter
                data = [x for x in data if not pd.isnull(x)]
                ws_sheet.write(header_cell, "", format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_lightblue_black)
                if column_letter == "A":
                    ws_sheet.set_column(all_column, 23)
                else:
                    ws_sheet.set_column(all_column, 23, None,
                                        {'level': 1, 'hidden': True})  # ,None,{'level': 1, 'hidden': True})

            elif "LSBP" in header or "Wellington" in header:
                kpis = df_total['index']
                ws_sheet.write(header_cell, header, format_darkblue_white)
                data = [x for x in data if not x == ""]

                for i in range(len(data)):
                    cell = column_letter + str(start_row_data + i)
                    value = data[i]
                    kpi = kpis[i]

                    if "%" in value:
                        value = float(value[:-1]) / 100
                        ws_sheet.write_number(cell, value, format_percentage)
                        if not "PR (%)" in kpi:
                            if "Availability" in kpi:
                                ws_sheet.conditional_format(cell, {'type': 'icon_set', 'icon_style': '3_traffic_lights',
                                                                   'icons': [{'criteria': '>=', 'type': 'number',
                                                                              'value': 0.97},
                                                                             {'criteria': '<', 'type': 'number',
                                                                              'value': 0.10},
                                                                             {'criteria': '<=', 'type': 'number',
                                                                              'value': 0.10}]})

                            else:
                                ws_sheet.conditional_format(cell, {'type': 'icon_set', 'icon_style': '3_traffic_lights',
                                                                   'icons': [
                                                                       {'criteria': '<=', 'type': 'number', 'value': 0},
                                                                       {'criteria': '>', 'type': 'number', 'value': 0},
                                                                       {'criteria': '>=', 'type': 'number',
                                                                        'value': 0.05}]})


                    else:
                        value = float(value.replace(",", ""))
                        ws_sheet.write_number(cell, value, format_number)

                ws_sheet.set_column(all_column, 16, None)


            elif "Time" in header:
                data = [x for x in data if not pd.isnull(x)]
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_day_hour)
                ws_sheet.set_column(all_column, 20, None, {'level': 1, 'hidden': True})

            elif "%" in header:
                to_collapse_column2 = column_letter
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_percentage)
                ws_sheet.set_column(all_column, 15, None, {'level': 1, 'hidden': True})

            elif "Capacity" in header or "(" in header:
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_number)
                ws_sheet.set_column(all_column, 15, None, {'level': 1, 'hidden': True})

            elif "ID" in header:
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_string_bold_wrapped)
                ws_sheet.set_column(all_column, 18, None, {'level': 1, 'hidden': True})



            elif "Site Name" in header:
                data = list(range(50))
                ws_sheet.write(header_cell, "", format_all_white)
                ws_sheet.write_column(data_cell, data, format_all_white)
                ws_sheet.set_column(all_column, 1, None, {'level': 1, 'hidden': True})

            elif "Space" in header:
                to_collapse_column = column_letter
                data = list(range(50))
                ws_sheet.write(header_cell, "+", format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_all_white)
                ws_sheet.set_column(all_column, 2, None, {'collapsed': True})

            elif "Actions" in header:
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_string)  # format_string_wrapped
                ws_sheet.set_column(all_column, 55, None, {'level': 1, 'hidden': True})


            else:
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_string)
                ws_sheet.set_column(all_column, 18, None, {'level': 1, 'hidden': True})

        level = level + 1
        start_column = start_column + n_columns_total
    # </editor-fold>

    # <editor-fold desc="FMECA AUX sheet">
    start_row_index = 1
    start_column_index = 1
    start_column = openpyxl.utils.cell.get_column_letter(1)
    dict_fmeca_table_range = {}
    for name, data in dict_fmeca_shapes.items():
        df = data[0]
        shape = data[1]

        n_row = shape[0]
        n_column = shape[1]

        start_column = openpyxl.utils.cell.get_column_letter(1)
        end_column = openpyxl.utils.cell.get_column_letter(shape[1])
        end_row = start_row_index + n_row

        start_cell = start_column + str(start_row_index)
        table_range = "$" + start_column + "$" + str(start_row_index + 1) + ":$" + end_column + "$" + str(end_row)
        dict_fmeca_table_range[name] = table_range

        # range
        # print(df)

        df.to_excel(writer, sheet_name='FMECA_AUX', startrow=start_row_index - 1, startcol=start_column_index - 1,
                    index=False)

        for i in range(len(df.columns)):
            range_name = df.columns[i]
            # print(range_name)
            column = openpyxl.utils.cell.get_column_letter(i + 1)
            range_cells = '$' + column + "$" + str(start_row_index + 1) + ":$" + column + "$" + str(end_row)
            workbook.define_name(range_name, '=FMECA_AUX!' + range_cells)
            """if "ategory" not in name:
                workbook.define_name(range_name, '=FMECA_AUX!' + range_cells)"""

        # Prepare next iteration
        start_row_index = start_row_index + n_row + 2
    # </editor-fold>

    # <editor-fold desc="Events' sheets">
    fmeca_columns = final_df_to_add['FMECA'].columns.to_list()
    n_rows_fmeca = final_df_to_add['FMECA'].shape[0]
    n_columns_fmeca = final_df_to_add['FMECA'].shape[1]
    reference_column = openpyxl.utils.cell.get_column_letter(
        final_df_to_add['FMECA'].columns.to_list().index('Fault') + 1)

    for sheet in final_df_to_add.keys():
        df = final_df_to_add[sheet]
        width = get_col_widths(df)
        n_rows = df.shape[0]
        n_columns = df.shape[1]
        try:
            ws_sheet = workbook.add_worksheet(sheet)
        except (xlsxwriter.exceptions.DuplicateWorksheetName, NameError):
            sheet = sheet + "_new"
            ws_sheet = workbook.add_worksheet(sheet)
        if "Closed" in sheet or "Active" in sheet:
            for i in range(len(df.columns)):
                header = df.columns[i]
                column_letter = openpyxl.utils.cell.get_column_letter(i + 1)
                header_cell = column_letter + '1'
                data_cell = column_letter + '2'
                all_column = column_letter + ':' + column_letter
                data = df[header].fillna("")

                if header == 'ID':
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_first_column)
                    ws_sheet.set_column(all_column, 18)

                elif "Time" in header:
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_day_hour)
                    ws_sheet.set_column(all_column, 19)

                elif "Capacity" in header or "(" in header:
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_number)
                    ws_sheet.set_column(all_column, 12)


                elif "Fa" in header or "ategory" in header or "Excludable" in header:
                    if header == "Resolution Category":
                        ws_sheet.write(header_cell, header, format_header)
                        ws_sheet.write_column(data_cell, data, format_string_unlocked)
                        ws_sheet.set_column(all_column, width[i + 1], unlocked)
                        ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                                 {'validate': 'list', 'source': ['Repair',
                                                                                 'Reset',
                                                                                 'Part Replacement',
                                                                                 'Unit Replacement']})

                    elif "Excludable" in header:
                        ws_sheet.write(header_cell, header, format_header)
                        ws_sheet.write_column(data_cell, data, format_string_unlocked)
                        ws_sheet.set_column(all_column, width[i + 1], unlocked)
                        # ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                        #  {'validate': 'list',
                        #   'source': ['OMC', 'Force Majeure', 'Curtailment', "N/A"]})

                    else:
                        fmeca_column_match = openpyxl.utils.cell.get_column_letter(fmeca_columns.index(header) + 1)
                        ws_sheet.write(header_cell, header, format_header)
                        ws_sheet.write_column(data_cell, data, format_string_unlocked)
                        ws_sheet.set_column(all_column, width[i + 1], unlocked)

                        # Add Data validation
                        if header == 'Fault':
                            ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                                     {'validate': 'list',
                                                      'source': '=FMECA_AUX!' + str(dict_fmeca_table_range['Faults'])})
                            fault_cell = data_cell

                        elif header == 'Fault Component':
                            ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                                     {'validate': 'list',
                                                      'source': '=INDIRECT(SUBSTITUTE(SUBSTITUTE(' +
                                                                fault_cell + ', " ", "_"), "-","_"))'})
                            fcomp_cell = data_cell

                        elif header == 'Failure Mode':
                            ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                                     {'validate': 'list',
                                                      'source': '=INDIRECT(SUBSTITUTE(SUBSTITUTE(' +
                                                                fault_cell + '&"_"&' +
                                                                fcomp_cell + '," ", "_"),"-","_"))'})
                            fmode_cell = data_cell

                        elif header == 'Failure Mechanism':
                            ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                                     {'validate': 'list',
                                                      'source': '=INDIRECT(SUBSTITUTE(SUBSTITUTE(' +
                                                                fault_cell + '&"_"&' +
                                                                fcomp_cell + '&"_"&' +
                                                                fmode_cell + ', " ", "_"), "-","_"))'})
                            fmec_cell = data_cell

                        elif header == 'Category':
                            ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                                     {'validate': 'list',
                                                      'source': '=INDIRECT(SUBSTITUTE(SUBSTITUTE(' +
                                                                fault_cell + '&"_"&' +
                                                                fcomp_cell + '&"_"&' +
                                                                fmode_cell + '&"_"&' +
                                                                fmec_cell + ', " ", "_"), "-","_"))'})
                            cat_cell = data_cell
                        elif header == 'Subcategory':
                            ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                                     {'validate': 'list',
                                                      'source': '=INDIRECT(SUBSTITUTE(SUBSTITUTE(' +
                                                                fault_cell + '&"_"&' +
                                                                fcomp_cell + '&"_"&' +
                                                                fmode_cell + '&"_"&' +
                                                                fmec_cell + '&"_"&' +
                                                                cat_cell + ', " ", "_"), "-","_"))'})
                            subcat_cell = data_cell

                elif header == "Incident Status":
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_string_unlocked)
                    ws_sheet.set_column(all_column, width[i + 1], unlocked)
                    ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                             {'validate': 'list', 'source': ['Open', 'Closed']})

                elif header == "Categorization Status":
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_string_unlocked)
                    ws_sheet.set_column(all_column, width[i + 1], unlocked)
                    ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                             {'validate': 'list', 'source': ['Pending', 'Completed']})

                elif header == 'Remediation' or header == 'Comments':
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_string_wrapped)
                    ws_sheet.set_column(all_column, 60)


                else:
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_string)
                    ws_sheet.set_column(all_column, width[i + 1])
        else:
            for i in range(len(df.columns)):
                header = df.columns[i]
                column_letter = openpyxl.utils.cell.get_column_letter(i + 1)
                header_cell = column_letter + '1'
                data_cell = column_letter + '2'
                all_column = column_letter + ':' + column_letter
                data = df[header].fillna("")

                if "ID" in header:
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_first_column)
                    ws_sheet.set_column(all_column, width[i + 1])
                else:
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_string)
                    ws_sheet.set_column(all_column, width[i + 1])

        ws_sheet.set_default_row(30)
    # </editor-fold>

    ws_active = workbook.get_worksheet_by_name("MTD Performance Overview")
    ws_active.activate()

    ws_fmeca_aux = workbook.get_worksheet_by_name('FMECA_AUX')
    ws_fmeca_aux.hide()

    writer.save()
    print('Done')

    return


def create_underperformance_report(underperformance_dest_file, incidents_corrected_period, performance_fleet_period,
                                   site_list):
    writer_und = pd.ExcelWriter(underperformance_dest_file, engine='xlsxwriter',
                                engine_kwargs={'options': {'nan_inf_to_errors': True}})
    workbook = writer_und.book

    # <editor-fold desc="Formats">
    # Format column header
    format_darkblue_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#002060', 'font_color': '#FFFFFF'})
    format_darkblue_white.set_bold()
    format_darkblue_white.set_text_wrap()

    format_lightblue_black = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'font_color': '#000000'})
    format_lightblue_black.set_bold()
    format_lightblue_black.set_text_wrap()
    format_lightblue_black.set_border()

    format_header = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#D9D9D9', 'font_color': '#000000'})
    format_header.set_bold()
    format_header.set_text_wrap()

    format_all_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFFFFF', 'font_color': '#FFFFFF'})
    format_all_black = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#000000', 'font_color': '#000000'})
    format_black_on_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#000000', 'font_color': '#FFFFFF'})

    # Format of specific column data
    format_day_data = workbook.add_format({'num_format': 'dd/mm/yyyy', 'valign': 'vcenter'})
    format_day_data.set_align('right')
    format_day_data.set_border()

    format_hour_data = workbook.add_format({'num_format': 'hh:mm:ss', 'valign': 'vcenter'})
    format_hour_data.set_align('right')
    format_hour_data.set_border()

    format_day_hour = workbook.add_format({'num_format': 'dd/mm/yyyy hh:mm:ss', 'valign': 'vcenter'})
    format_day_hour.set_align('right')
    format_day_hour.set_border()

    # Format numbers
    format_number = workbook.add_format({'num_format': '#,##0.00', 'align': 'center', 'valign': 'vcenter'})
    format_number.set_border()

    format_nodecimal = workbook.add_format({'num_format': '0', 'align': 'center', 'valign': 'vcenter'})
    format_nodecimal.set_border()

    format_percentage = workbook.add_format({'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter'})
    format_percentage.set_border()

    format_percentage_good = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#C6EFCE',
         'font_color': '#006100'})
    format_percentage_good.set_border()
    format_percentage_mid = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFEB9C',
         'font_color': '#9C5700'})
    format_percentage_mid.set_border()
    format_percentage_bad = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFC7CE',
         'font_color': '#9C0006'})
    format_percentage_bad.set_border()

    # Format strings
    format_string = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
    format_string.set_border()

    format_string_wrapped = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
    format_string_wrapped.set_text_wrap()
    format_string_wrapped.set_border()

    format_string_unlocked = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'locked': False})
    unlocked = workbook.add_format({'locked': False})
    format_string_unlocked.set_border()

    format_string_bold = workbook.add_format({'align': 'right', 'valign': 'vcenter'})
    format_string_bold.set_bold()
    format_string_bold.set_border()

    format_string_bold_wrapped = workbook.add_format({'align': 'right', 'valign': 'vcenter'})
    format_string_bold_wrapped.set_bold()
    format_string_bold_wrapped.set_border()
    format_string_bold_wrapped.set_text_wrap()

    format_first_column = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#F2F2F2', 'font_color': '#000000'})
    format_first_column.set_bold()
    format_first_column.set_border()
    format_first_column.set_text_wrap()
    # </editor-fold>

    # <editor-fold desc="Performance Overview Sheet">
    sheet = "Performance Overview"
    try:
        ws_sheet = workbook.add_worksheet(sheet)
    except (xlsxwriter.exceptions.DuplicateWorksheetName, NameError):
        sheet = sheet + "_new"
        ws_sheet = workbook.add_worksheet(sheet)

    df_performance = performance_fleet_period.T


    sites = list(df_performance.columns)

    start_row_header = 1
    start_row_data = 2
    start_column = 0

    for site in sites:
        level = 0
        start_row_header_str = str(start_row_header)
        start_row_data_str = str(start_row_data)

        performance_site = df_performance.loc[:, [site]].reset_index()
        n_rows_performance = performance_site.shape[0] + 1
        n_columns_performance = performance_site.shape[1]

        df_total = performance_site

        max_rows = n_rows_performance
        n_columns_total = df_total.shape[1]

        width = get_col_widths(df_total)

        print("\n", df_total)

        for i in range(start_column, start_column + n_columns_total):

            header = df_total.columns[i - start_column]

            column_letter = openpyxl.utils.cell.get_column_letter(i + 1)
            header_cell = column_letter + start_row_header_str
            data_cell = column_letter + start_row_data_str
            all_column = column_letter + ':' + column_letter

            # print('Header: ', header, "\n", 'Header cell:', header_cell, "\n", "Data Cell: ", data_cell ,"\n")

            data = list(df_total[header].fillna(""))

            if header == "index":
                to_collapse_column1 = column_letter
                data = [x for x in data if not pd.isnull(x)]
                ws_sheet.write(header_cell, "", format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_lightblue_black)
                if column_letter == "A":
                    ws_sheet.set_column(all_column, 23)
                else:
                    ws_sheet.set_column(all_column, 0)  # ,None,{'level': 1, 'hidden': True})

            elif header in site_list:
                kpis = df_total['index']
                ws_sheet.write(header_cell, header, format_darkblue_white)
                data = [x for x in data if not x == ""]

                for i in range(len(data)):
                    cell = column_letter + str(start_row_data + i)
                    value = data[i]
                    kpi = kpis[i]

                    if "%" in value:
                        value = float(value[:-1]) / 100
                        ws_sheet.write_number(cell, value, format_percentage)
                        if not "PR (%)" in kpi:
                            if "Availability" in kpi:
                                ws_sheet.conditional_format(cell, {'type': 'icon_set', 'icon_style': '3_traffic_lights',
                                                                   'icons': [{'criteria': '>=', 'type': 'number',
                                                                              'value': 0.97},
                                                                             {'criteria': '<', 'type': 'number',
                                                                              'value': 0.10},
                                                                             {'criteria': '<=', 'type': 'number',
                                                                              'value': 0.10}]})

                            else:
                                ws_sheet.conditional_format(cell, {'type': 'icon_set', 'icon_style': '3_traffic_lights',
                                                                   'icons': [
                                                                       {'criteria': '<=', 'type': 'number', 'value': 0},
                                                                       {'criteria': '>', 'type': 'number', 'value': 0},
                                                                       {'criteria': '>=', 'type': 'number',
                                                                        'value': 0.05}]})


                    else:
                        value = float(value.replace(",", ""))
                        ws_sheet.write_number(cell, value, format_number)

                ws_sheet.set_column(all_column, 16, None)


            else:
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_string)
                ws_sheet.set_column(all_column, 18, None)  # ,{'level': 1, 'hidden': True})

        level = level + 1
        start_column = start_column + n_columns_total
    # </editor-fold>

    ws_active = workbook.get_worksheet_by_name("Performance Overview")
    ws_active.activate()

    incidents_corrected_period.to_excel(writer_und, sheet_name='Underperformance Report', index=False)

    writer_und.save()

    print('Done')

    return

# </editor-fold>
